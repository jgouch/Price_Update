import pandas as pd
import re
import os
import sys
import warnings
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from typing import Optional, Dict

# --- 1. SETUP ---
warnings.simplefilter(action='ignore', category=FutureWarning)

# --- 2. CLEANING LOGIC ---
def clean_sheet_name_specific(sheet_name):
    name = re.sub(r'^\d+_', '', sheet_name)
    for word in ['Mausoleum', 'Bldg', 'Building']:
        name = name.replace(word, '')
    return name.strip()

def clean_sheet_name_generic(sheet_name):
    name = clean_sheet_name_specific(sheet_name)
    for word in ['Columbarium', 'Niches', 'Garden']:
        name = name.replace(word, '')
    return name.strip()

def clean_row_name(row_str):
    s = str(row_str).strip().upper()
    if 'ALL LEVEL' in s: return 'ALL'
    s = re.sub(r'\(.*?\)', '', s)
    s = s.replace('UNCOVERED', '').replace('COVERED', '').replace('ELEVATION', '')
    s = s.replace('—', '-').replace('–', '-')
    if '-' in s: s = s.split('-')[0]
    return s.strip()

# --- 3. MAPPING ---
def identify_columns(df):
    cols = [str(c) for c in df.columns]
    mapping = {'Garden': None, 'Row': None, 'Status': None}
    
    for c in cols:
        if 'GARDEN' in c.upper() or 'GROUP' in c.upper() or 'LOCATION' in c.upper():
            mapping['Garden'] = c; break
    
    candidates = [c for c in cols if any(x in c.upper() for x in ['SECTION', 'ROW', 'BLOCK', 'LOT', 'TIER'])]
    if candidates:
        best = next((x for x in candidates if 'SECTION' in x.upper()), None)
        mapping['Row'] = best if best else candidates[0]

    for c in cols:
        if 'STATUS' in c.upper() or 'STATE' in c.upper():
            mapping['Status'] = c; break
            
    return mapping

def validate_column_mapping(col_map):
    missing = [key for key, value in col_map.items() if value is None]
    if missing:
        raise ValueError(
            "Missing required inventory columns: "
            f"{', '.join(missing)}. Please verify the inventory headers."
        )

def is_blank_string(value: Optional[str]) -> bool:
    return not str(value or "").strip()

def garden_exists_in_inventory(df_inventory, garden_name, col_map):
    col_garden = col_map['Garden']
    mask = df_inventory[col_garden].astype(str).str.contains(garden_name, case=False, na=False)
    return mask.any()

# --- 4. CALCULATIONS ---
def calculate_percent_sold(df_inventory, garden_name_full, col_map):
    col_garden, col_section, col_status = col_map['Garden'], col_map['Row'], col_map['Status']
    status_avail = ['Available', 'Serviceable', 'For Sale', 'Vacant']
    
    if is_blank_string(garden_name_full):
        return None

    parts = re.split(r'[-–]', garden_name_full)
    main_garden = parts[0].strip()
    sub_section = parts[1].strip() if len(parts) > 1 else None

    if is_blank_string(main_garden):
        return None

    garden_mask = df_inventory[col_garden].astype(str).str.contains(main_garden, case=False, na=False)
    garden_data = df_inventory[garden_mask]
    
    if sub_section and not garden_data.empty:
        section_mask = garden_data[col_section].astype(str).str.contains(sub_section, case=False, na=False)
        if section_mask.any(): garden_data = garden_data[section_mask]
    
    total = len(garden_data)
    if total == 0: return None
        
    avail_mask = garden_data[col_status].astype(str).str.contains('|'.join(status_avail), case=False, na=False)
    return (total - len(garden_data[avail_mask])) / total

def count_row_availability(df_inventory, garden_name, row_name, col_map):
    col_garden, col_row, col_status = col_map['Garden'], col_map['Row'], col_map['Status']
    status_avail = ['Available', 'Serviceable', 'For Sale', 'Vacant']

    if is_blank_string(garden_name):
        return "N/A"

    garden_mask = df_inventory[col_garden].astype(str).str.contains(garden_name, case=False, na=False)
    garden_data = df_inventory[garden_mask]
    
    if garden_data.empty: return "N/A"

    target = clean_row_name(row_name)
    if target == 'ALL': row_data = garden_data
    else:
        row_mask = garden_data[col_row].astype(str).apply(clean_row_name) == target
        row_data = garden_data[row_mask]
    
    if len(row_data) == 0: return None
    
    avail_mask = row_data[col_status].astype(str).str.contains('|'.join(status_avail), case=False, na=False)
    return len(row_data[avail_mask])

def coerce_numeric_cell(value: object) -> Optional[float]:
    if value is None:
        return None
    raw = str(value).strip()
    if not raw:
        return None
    cleaned = raw.replace(",", "")
    if re.fullmatch(r"-?\d+(\.\d+)?", cleaned):
        return float(cleaned)
    return None

# --- 5. FORMATTING (FIXED) ---
def apply_professional_formatting(file_path):
    print("\n🎨 Applying Final Polish (removing Unnamed columns, adding $, %)...")
    
    wb = load_workbook(file_path)
    
    # Styles
    HEADER_COLOR = "363636" # Dark Grey
    HEADER_TEXT = "FFFFFF"  # White
    ROW_ALT_COLOR = "F2F2F2"
    FONT_NAME = "Calibri"
    
    header_font = Font(name=FONT_NAME, size=11, bold=True, color=HEADER_TEXT)
    header_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin', color="D9D9D9"), right=Side(style='thin', color="D9D9D9"), top=Side(style='thin', color="D9D9D9"), bottom=Side(style='thin', color="D9D9D9"))

    for ws in wb.worksheets:
        # 1. FIX HEADERS (Remove empty rows above real headers)
        # Find the real header row (first row with "Price" or "Level" or "Garden")
        header_row_idx = 1
        for row in ws.iter_rows(min_row=1, max_row=10):
            values = [str(c.value).upper() for c in row if c.value]
            if any(x in str(values) for x in ['PRICE', 'GARDEN', 'LEVEL', 'ROW', 'SECTION']):
                header_row_idx = row[0].row
                break
        
        # If headers are not in row 1, delete rows above
        if header_row_idx > 1:
            ws.delete_rows(1, header_row_idx - 1)

        # 2. FORMAT CELLS
        for col in ws.columns:
            # Auto-Fit Width
            max_len = 0
            col_letter = col[0].column_letter
            header_val = str(ws[f"{col_letter}1"].value).upper()
            
            for cell in col:
                # Value Checking
                if cell.value is not None:
                    val_str = str(cell.value)
                    max_len = max(max_len, len(val_str))
                    
                    # Convert Text Numbers to Real Numbers (Fixes Green Triangles)
                    if cell.row > 1:
                        coerced = coerce_numeric_cell(cell.value)
                        if coerced is not None:
                            cell.value = coerced
                
                # Apply Styles
                cell.font = Font(name=FONT_NAME, size=10)
                cell.border = thin_border
                
                if cell.row == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                else:
                    if cell.row % 2 == 0:
                        cell.fill = PatternFill(start_color=ROW_ALT_COLOR, end_color=ROW_ALT_COLOR, fill_type="solid")
                    
                    # Number Formatting & Centering
                    if 'PRICE' in header_val or 'TOTAL' in header_val or 'RIGHT' in header_val:
                        cell.number_format = '"$"#,##0'
                        cell.alignment = center_align
                    elif '%' in header_val or 'SOLD' in header_val:
                        cell.number_format = '0%'
                        cell.alignment = center_align
                    elif 'QTY' in header_val or 'AVAIL' in header_val:
                        cell.alignment = center_align
                        if str(cell.value).upper() == "SOLD OUT":
                            cell.font = Font(name=FONT_NAME, size=10, color="FF0000", bold=True)
                            
            ws.column_dimensions[col_letter].width = min((max_len + 2) * 1.1, 50)

    wb.save(file_path)

# --- 6. MAIN ---
def main():
    if len(sys.argv) < 3:
        print("Usage: python3 update_v12.py [File1] [File2]")
        return

    path_a = sys.argv[1].strip().replace("'", "").replace('"', "")
    path_b = sys.argv[2].strip().replace("'", "").replace('"', "")

    f1, f2 = os.path.basename(path_a).lower(), os.path.basename(path_b).lower()
    if 'inventory' in f1: inv_path, master_path = path_a, path_b
    elif 'inventory' in f2: inv_path, master_path = path_b, path_a
    else: print("❌ Error: One file must include 'Inventory' in the name."); return

    print(f"\n📂 Inventory: {os.path.basename(inv_path)}")
    print(f"📘 Master Book: {os.path.basename(master_path)}")

    # READ & UPDATE
    try:
        # Load Inventory (Header Row 3)
        df_inv = pd.read_excel(inv_path, header=2)
        col_map = identify_columns(df_inv)
        validate_column_mapping(col_map)
        
        # Load Master (Detect Headers per sheet later)
        sheet_dict = pd.read_excel(master_path, sheet_name=None, header=None) # Read raw to handle variable headers
    except Exception as e: print(f"❌ Error: {e}"); return

    print("⚙️  Processing Updates...")
    folder = os.path.dirname(master_path)
    output_path = os.path.join(folder, 'Harpeth_Hills_Price_Book_FINAL.xlsx')

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for sheet_name, df_raw in sheet_dict.items():
            
            # Find Header Row dynamically for processing
            header_idx = 0
            for i, row in df_raw.head(10).iterrows():
                vals = [str(x).upper() for x in row.values]
                if any(k in str(vals) for k in ['PRICE', 'GARDEN', 'LEVEL', 'ROW']):
                    header_idx = i; break
            
            # Reload sheet with correct header
            df = pd.read_excel(master_path, sheet_name=sheet_name, header=header_idx)
            
            # LOGIC
            search_name_spec = clean_sheet_name_specific(sheet_name)
            search_name_gen = clean_sheet_name_generic(sheet_name)
            final_search_name = search_name_spec if garden_exists_in_inventory(df_inv, search_name_spec, col_map) else search_name_gen
            if is_blank_string(final_search_name):
                print(f"⚠️  Skipping sheet '{sheet_name}' due to empty garden name after cleanup.")
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                continue
            
            cols = [str(c).upper() for c in df.columns]
            
            # % Sold
            percent_headers = [c for c in cols if '%' in c or 'PERCENT' in c]
            if percent_headers and 'GARDEN' in cols:
                g_col = next(c for c in df.columns if str(c).upper() == 'GARDEN')
                s_col = next(
                    c for c in df.columns
                    if ('%' in str(c)) or ('PERCENT' in str(c).upper())
                )
                for idx, row in df.iterrows():
                    val = calculate_percent_sold(df_inv, str(row[g_col]), col_map)
                    if val is not None: df.at[idx, s_col] = val

            # Counts
            r_cands = [c for c in df.columns if any(x in str(c).upper() for x in ['ROW', 'LEVEL', 'SECTION', 'STATION', 'PRODUCT'])]
            q_cands = [c for c in df.columns if any(x in str(c).upper() for x in ['AVAIL', 'QTY', 'STATUS'])]
            
            if r_cands and q_cands:
                r_col, q_col = r_cands[0], q_cands[0]
                for idx, row in df.iterrows():
                    val = count_row_availability(df_inv, final_search_name, str(row[r_col]), col_map)
                    if val is not None and val != "N/A":
                        df.at[idx, q_col] = "Sold Out" if val == 0 else val

            df.to_excel(writer, sheet_name=sheet_name, index=False)

    apply_professional_formatting(output_path)
    print(f"\nSUCCESS! 🚀\nFile saved: {output_path}")

if __name__ == "__main__":
    main()
