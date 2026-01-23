import pandas as pd
import re
import os
import sys
import warnings
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- 1. SETUP ---
warnings.simplefilter(action='ignore', category=FutureWarning)

# --- 2. INTELLIGENT HEADER SCANNER ---
def find_inventory_header(file_path):
    """
    Scans the first 20 rows of the inventory file to find the row 
    that contains column headers like 'Garden', 'Status', or 'Property'.
    """
    print("🔎 Scanning Inventory file for headers...")
    
    # Read first 20 rows without a header
    df_temp = pd.read_excel(file_path, header=None, nrows=20)
    
    best_row = 0
    max_matches = 0
    
    # Keywords to look for in the header row
    keywords = ['GARDEN', 'STATUS', 'STATE', 'PROPERTY', 'LOCATION', 'DESCRIPTION', 'SPACE', 'LOT']
    
    for i, row in df_temp.iterrows():
        # Convert row values to string and upper case
        row_str = " ".join([str(x).upper() for x in row.values if pd.notna(x)])
        
        # Count how many keywords appear in this row
        matches = sum(1 for k in keywords if k in row_str)
        
        if matches > max_matches:
            max_matches = matches
            best_row = i
            
    if max_matches >= 2:
        print(f"✅ Found Inventory Headers at Row {best_row + 1}")
        return best_row
    else:
        print("⚠️  Could not confidently find headers. Defaulting to Row 3 (Index 2).")
        return 2

# --- 3. CLEANING FUNCTIONS ---
def clean_sheet_name_specific(sheet_name):
    name = re.sub(r'^\d+_', '', sheet_name)
    for word in ['Mausoleum', 'Bldg', 'Building']:
        name = name.replace(word, '')
    return name.strip()

def clean_sheet_name_generic(sheet_name):
    name = clean_sheet_name_specific(sheet_name)
    for word in ['Columbarium', 'Niches', 'Garden']:
        name = name.replace(word, '')
    return name.strip()

def clean_row_name(row_str):
    s = str(row_str).strip().upper()
    if 'ALL LEVEL' in s: return 'ALL'
    s = re.sub(r'\(.*?\)', '', s)
    s = s.replace('UNCOVERED', '').replace('COVERED', '').replace('ELEVATION', '')
    s = s.replace('—', '-').replace('–', '-')
    if '-' in s: s = s.split('-')[0]
    return s.strip()

# --- 4. MAPPING & CHECKING ---
def identify_columns(df):
    cols = [str(c) for c in df.columns]
    mapping = {'Garden': None, 'Row': None, 'Status': None}
    
    # GARDEN / PROPERTY ID
    for c in cols:
        cu = c.upper()
        if 'GARDEN' in cu or 'GROUP' in cu or 'LOCATION' in cu or 'PROPERTY' in cu:
            mapping['Garden'] = c; break
            
    # ROW / SECTION
    candidates = [c for c in cols if any(x in c.upper() for x in ['SECTION', 'ROW', 'BLOCK', 'LOT', 'TIER'])]
    if candidates:
        # Prefer "Section" or "Row"
        best = next((x for x in candidates if 'SECTION' in x.upper()), None)
        if not best: best = next((x for x in candidates if 'ROW' in x.upper()), None)
        mapping['Row'] = best if best else candidates[0]

    # STATUS
    for c in cols:
        if 'STATUS' in c.upper() or 'STATE' in c.upper():
            mapping['Status'] = c; break
            
    return mapping

def garden_exists_in_inventory(df_inventory, garden_name, col_map):
    col_garden = col_map['Garden']
    if not col_garden: return False
    mask = df_inventory[col_garden].astype(str).str.contains(garden_name, case=False, na=False, regex=False)
    return mask.any()

# --- 5. CALCULATIONS ---
def calculate_percent_sold(df_inventory, garden_name_full, col_map):
    col_garden, col_section, col_status = col_map['Garden'], col_map['Row'], col_map['Status']
    status_avail = ['Available', 'Serviceable', 'For Sale', 'Vacant']
    
    # 1. SPLIT & CLEAN
    parts = re.split(r'[-–]', garden_name_full)
    main_garden = parts[0].split('(')[0].strip() # Clean parentheses
    sub_section = parts[1].strip() if len(parts) > 1 else None

    # 2. FILTER
    garden_mask = df_inventory[col_garden].astype(str).str.contains(main_garden, case=False, na=False, regex=False)
    garden_data = df_inventory[garden_mask]
    
    if sub_section and not garden_data.empty:
        section_mask = garden_data[col_section].astype(str).str.contains(sub_section, case=False, na=False, regex=False)
        if section_mask.any(): garden_data = garden_data[section_mask]
    
    total = len(garden_data)
    
    # --- GRACE DIAGNOSTIC ---
    if 'Grace' in garden_name_full:
        print(f"   > Checking '{garden_name_full}' -> Cleaned: '{main_garden}' -> Found: {total} matches")

    if total == 0: return None
        
    avail_mask = garden_data[col_status].astype(str).str.contains('|'.join(status_avail), case=False, na=False)
    return (total - len(garden_data[avail_mask])) / total

def count_row_availability(df_inventory, garden_name, row_name, col_map):
    col_garden, col_row, col_status = col_map['Garden'], col_map['Row'], col_map['Status']
    status_avail = ['Available', 'Serviceable', 'For Sale', 'Vacant']

    garden_mask = df_inventory[col_garden].astype(str).str.contains(garden_name, case=False, na=False, regex=False)
    garden_data = df_inventory[garden_mask]
    
    if garden_data.empty: return "N/A"

    target = clean_row_name(row_name)
    if target == 'ALL': row_data = garden_data
    else:
        row_mask = garden_data[col_row].astype(str).apply(clean_row_name) == target
        row_data = garden_data[row_mask]
    
    if len(row_data) == 0: return None
    
    avail_mask = row_data[col_status].astype(str).str.contains('|'.join(status_avail), case=False, na=False)
    return len(row_data[avail_mask])

# --- 6. SURGICAL UPDATE ---
def surgical_update(inv_path, master_path, output_path):
    print("⚙️  Loading files...")
    
    # A. READ INVENTORY (With Auto-Header Detect)
    try:
        header_row = find_inventory_header(inv_path)
        df_inv = pd.read_excel(inv_path, header=header_row)
        col_map = identify_columns(df_inv)
        
        print(f"   > Mapped Columns: {col_map}")
        
        # Diagnostic Scan for "Grace"
        if col_map['Garden']:
            grace_count = df_inv[col_map['Garden']].astype(str).str.contains('Grace', case=False).sum()
            print(f"   > DIAGNOSTIC: Found 'Grace' in inventory {grace_count} times.")
        else:
            print("   > ❌ CRITICAL ERROR: Could not identify 'Garden' column in inventory.")
            
    except Exception as e: print(f"❌ Inventory Error: {e}"); return

    # B. LOAD MASTER
    wb = load_workbook(master_path)
    
    # Styles
    HEADER_COLOR = "363636"
    HEADER_TEXT = "FFFFFF"
    ROW_ALT_COLOR = "F2F2F2"
    FONT_NAME = "Calibri"
    
    header_font = Font(name=FONT_NAME, size=11, bold=True, color=HEADER_TEXT)
    header_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
    alt_fill = PatternFill(start_color=ROW_ALT_COLOR, end_color=ROW_ALT_COLOR, fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin', color="D9D9D9"), right=Side(style='thin', color="D9D9D9"), top=Side(style='thin', color="D9D9D9"), bottom=Side(style='thin', color="D9D9D9"))

    print("⚙️  Processing Worksheets...")
    for ws in wb.worksheets:
        header_row = 1
        col_indices = {}
        
        # 1. FIND MASTER SHEET HEADER
        for r in range(1, 21):
            row_vals = [str(ws.cell(row=r, column=c).value).upper() for c in range(1, ws.max_column + 1)]
            if any(x in str(row_vals) for x in ['PRICE', 'GARDEN', 'LEVEL', 'ROW', 'SECTION', 'AVAIL']):
                header_row = r
                for c in range(1, ws.max_column + 1):
                    val = str(ws.cell(row=r, column=c).value).upper()
                    if val != 'NONE': col_indices[c] = val
                break
        
        sheet_name = ws.title
        search_name_spec = clean_sheet_name_specific(sheet_name)
        search_name_gen = clean_sheet_name_generic(sheet_name)
        
        # Determine strictness of search
        if garden_exists_in_inventory(df_inv, search_name_spec, col_map):
            final_search_name = search_name_spec
        else:
            final_search_name = search_name_gen

        # 2. UPDATE ROWS
        data_row_count = 0
        
        for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row):
            r_idx = row[0].row
            row_name = None
            garden_name = None
            is_secondary_header = False
            
            # Identify row types
            for cell in row:
                if cell.column in col_indices:
                    col_name = col_indices[cell.column]
                    if any(x in col_name for x in ['ROW', 'LEVEL', 'SECTION', 'PRODUCT']):
                        if str(cell.value).strip().upper() in ['ROW', 'LEVEL', 'SECTION', 'PRODUCT', 'STATION', 'NICHE #']:
                            is_secondary_header = True

            is_header = (row[0].row == header_row) or is_secondary_header
            
            # --- FORMATTING & CLEANING ---
            for cell in row:
                col_name = col_indices.get(cell.column, "")
                cell.border = thin_border
                
                if is_header:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                else:
                    cell.font = Font(name=FONT_NAME, size=10, color="000000")
                    if data_row_count % 2 == 0:
                        cell.fill = alt_fill
                    else:
                        cell.fill = PatternFill(fill_type=None)
                    
                    # Number Fixes (Green Triangle Killer)
                    if any(x in col_name for x in ['PRICE', 'TOTAL', 'RIGHT', 'COST', 'PLAQUE', 'ETCHING', 'FRONT', 'CRYPT', 'NICHE', 'MEMORIAL', 'BURIAL', 'OPTION']):
                        val_str = str(cell.value).upper()
                        if val_str not in ['SINGLE', 'COMPANION', 'TANDEM', 'SIDE-BY-SIDE', 'NONE', 'None', '']:
                            cell.number_format = '"$"#,##0'
                            cell.alignment = center_align
                            try:
                                if cell.value is not None:
                                    clean_val = str(cell.value).replace('$', '').replace(',', '').strip()
                                    if clean_val.replace('.', '', 1).isdigit():
                                        cell.value = float(clean_val)
                            except: pass

                    elif any(x in col_name for x in ['%', 'SOLD']) and 'QTY' not in col_name:
                        cell.number_format = '0%'
                        cell.alignment = center_align
                        
                    elif any(x in col_name for x in ['QTY', 'AVAIL', 'STATUS']):
                        cell.alignment = center_align
                        cell.number_format = '0' 
                        val_str = str(cell.value).upper()
                        if val_str == "SOLD OUT":
                            cell.font = Font(name=FONT_NAME, size=10, color="FF0000", bold=True)
                            cell.number_format = 'General'
                        else:
                            try:
                                if str(cell.value).replace('.', '', 1).isdigit():
                                    cell.value = int(float(str(cell.value)))
                                    if 0 < cell.value < 4:
                                        cell.font = Font(name=FONT_NAME, size=10, color="E26B0A", bold=True)
                            except: pass

                if not is_header:
                    val_str = str(cell.value).strip()
                    if any(x in col_name for x in ['ROW', 'LEVEL', 'SECTION', 'STATION', 'PRODUCT']):
                        row_name = val_str
                    if 'GARDEN' in col_name:
                        garden_name = val_str
            
            # --- DATA INJECTION ---
            if not is_header:
                # Update % Sold
                if garden_name:
                    new_pct = calculate_percent_sold(df_inv, garden_name, col_map)
                    if new_pct is not None:
                        for c, name in col_indices.items():
                            if '%' in name: ws.cell(row=r_idx, column=c).value = new_pct

                # Update Counts
                if row_name and row_name != 'None' and row_name != '':
                    count = count_row_availability(df_inv, final_search_name, row_name, col_map)
                    if count is not None and count != "N/A":
                        for c, name in col_indices.items():
                            if any(x in name for x in ['AVAIL', 'QTY', 'STATUS']):
                                cell = ws.cell(row=r_idx, column=c)
                                if count == 0:
                                    cell.value = "Sold Out"
                                    cell.font = Font(name=FONT_NAME, size=10, color="FF0000", bold=True)
                                    cell.number_format = 'General'
                                else:
                                    cell.value = int(count)
                                    cell.number_format = '0'
                                    if count < 4:
                                        cell.font = Font(name=FONT_NAME, size=10, color="E26B0A", bold=True)
                
                data_row_count += 1

        # 3. AUTO-FIT (Merged Cell Safe)
        for i, col in enumerate(ws.columns, 1):
            max_len = 0
            col_letter = get_column_letter(i)
            for cell in col:
                if cell.row >= header_row and cell.value:
                    try: max_len = max(max_len, len(str(cell.value)))
                    except: pass
            if max_len > 0:
                ws.column_dimensions[col_letter].width = min((max_len + 2) * 1.1, 50)

    wb.save(output_path)
    print(f"\nSUCCESS! 🚀\nFile saved: {output_path}")

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python3 update_v21.py [File1] [File2]")
    else:
        path_a = sys.argv[1].strip().replace("'", "").replace('"', "")
        path_b = sys.argv[2].strip().replace("'", "").replace('"', "")
        
        f1, f2 = os.path.basename(path_a).lower(), os.path.basename(path_b).lower()
        if 'inventory' in f1: inv_path, master_path = path_a, path_b
        elif 'inventory' in f2: inv_path, master_path = path_b, path_a
        else:
            print("❌ Error: One file must include 'Inventory' in the name.")
            sys.exit()

        folder = os.path.dirname(master_path)
        output_path = os.path.join(folder, 'Harpeth_Hills_Price_Book_FINAL.xlsx')
        surgical_update(inv_path, master_path, output_path)
