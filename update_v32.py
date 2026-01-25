import pandas as pd
import re
import os
import sys
import warnings
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

# --- 1. SETUP ---
warnings.simplefilter(action='ignore', category=FutureWarning)

# --- 2. HEADER SCANNER ---
def find_inventory_header(file_path):
    print("🔎 Scanning Inventory file for headers...")
    try:
        # Read first 20 rows to find the header
        df_temp = pd.read_excel(file_path, header=None, nrows=20)
    except Exception as e:
        print(f"❌ Error reading file: {e}")
        return 0

    best_row = 0
    max_matches = 0
    keywords = ['SECTION', 'SPACE', 'STATUS', 'COST', 'RIGHTS', 'GARDEN']
    
    for i, row in df_temp.iterrows():
        # Case-insensitive string match
        row_str = " ".join([str(x).upper() for x in row.values if pd.notna(x)])
        matches = sum(1 for k in keywords if k in row_str)
        if matches > max_matches:
            max_matches = matches
            best_row = i
            
    if max_matches >= 2:
        print(f"✅ Found Inventory Headers at Row {best_row + 1}")
        return best_row
    else:
        print("⚠️  Could not confidently find headers. Defaulting to Row 3.")
        return 2

# --- 3. CLEANING FUNCTIONS ---
def clean_sheet_name_specific(sheet_name):
    name = re.sub(r'^\d+_', '', sheet_name)
    for word in ['Mausoleum', 'Bldg', 'Building', 'Garden']:
        name = name.replace(word, '')
    return name.strip()

def clean_sheet_name_generic(sheet_name):
    name = clean_sheet_name_specific(sheet_name)
    for word in ['Columbarium', 'Niches']:
        name = name.replace(word, '')
    return name.strip()

def clean_row_name(row_str):
    s = str(row_str).strip().upper()
    if 'ALL LEVEL' in s: return 'ALL'
    s = re.sub(r'\(.*?\)', '', s)
    s = s.replace('UNCOVERED', '').replace('COVERED', '').replace('ELEVATION', '').replace('LEVEL', '')
    s = s.replace('—', '-').replace('–', '-')
    if '-' in s: s = s.split('-')[0]
    return s.strip()

def super_clean_name(name):
    """
    Aggressive cleaner. 
    Returns empty string if the result is too short to be a valid match.
    """
    if not isinstance(name, str): return ""
    name = re.sub(r'\(.*?\)', '', name)
    name = re.sub(r'\d+', '', name)
    for word in ['GARDEN', 'SECTION', 'LOC', 'LOCATION', 'BLOCK', 'OF', 'THE']:
        name = name.upper().replace(word, '')
    name = re.sub(r'[^\w\s]', '', name)
    cleaned = name.strip().upper()
    
    # SAFETY: If cleaned name is too short (e.g. "A" or ""), return empty to prevent over-matching
    if len(cleaned) < 2: return "" 
    return cleaned

# --- 4. ROBUST MAPPING ---
def identify_columns(df):
    # Convert actual columns to uppercase for safe matching
    cols_map = {str(c).strip().upper(): c for c in df.columns}
    mapping = {'Garden': None, 'Row': None, 'Status': None}
    
    # 1. GARDEN (Prioritize 'SECTION' or specific Garden names)
    if 'SECTION' in cols_map:
        mapping['Garden'] = cols_map['SECTION']
    else:
        for c_up, c_orig in cols_map.items():
            if 'GARDEN' in c_up or 'LOCATION' in c_up:
                mapping['Garden'] = c_orig; break

    # 2. ROW (Prioritize 'SPACE' -> 'LOT' -> 'ROW')
    if 'SPACE' in cols_map:
        mapping['Row'] = cols_map['SPACE']
    elif 'LOT' in cols_map:
        mapping['Row'] = cols_map['LOT']
    else:
        for c_up, c_orig in cols_map.items():
            if 'ROW' in c_up or 'TIER' in c_up:
                mapping['Row'] = c_orig; break

    # 3. STATUS
    for c_up, c_orig in cols_map.items():
        if 'STATUS' in c_up or 'STATE' in c_up:
            mapping['Status'] = c_orig; break
            
    return mapping

def garden_exists_in_inventory(df_inventory, garden_name, col_map):
    col_garden = col_map['Garden']
    if not col_garden: return False
    
    target = super_clean_name(garden_name)
    if not target: return False # Safety check for empty names
    
    inv_clean = df_inventory[col_garden].astype(str).apply(super_clean_name)
    return inv_clean.str.contains(target, case=False, na=False).any()

# --- 5. CALCULATIONS ---
def is_grace_sidewalk(space_str):
    """
    Grace Sidewalk/Roadside Sections: 
    30, 40, 50, 60-64, 70-74, 80, 90, 100, 110, 120
    """
    if not isinstance(space_str, str): return False
    match = re.search(r'Lot/Section\s+(\d+)', space_str, re.IGNORECASE)
    if not match: return False
    
    try:
        section_num = int(match.group(1))
        sidewalk_sections = [30, 40, 50, 80, 90, 100, 110, 120]
        sidewalk_sections.extend(range(60, 65)) # 60-64
        sidewalk_sections.extend(range(70, 75)) # 70-74
        return section_num in sidewalk_sections
    except:
        return False

def calculate_percent_sold(df_inventory, garden_name_full, col_map):
    col_garden, col_section, col_status = col_map['Garden'], col_map['Row'], col_map['Status']
    # Safeguard against missing columns
    if not col_garden or not col_status: return None

    status_avail = ['Available', 'Serviceable', 'For Sale', 'Vacant']
    parts = re.split(r'[-–]', garden_name_full)
    
    # Clean target name
    target_garden = super_clean_name(parts[0])
    if not target_garden: return None # Safety for empty names

    sub_section = parts[1].strip() if len(parts) > 1 else None

    # Filter by Garden
    inv_gardens = df_inventory[col_garden].astype(str).apply(super_clean_name)
    garden_mask = inv_gardens.str.contains(target_garden, case=False, na=False)
    garden_data = df_inventory[garden_mask]
    
    if garden_data.empty: return None

    # --- GRACE LOGIC ---
    if target_garden == "GRACE" and col_section:
        is_sidewalk = garden_data[col_section].apply(is_grace_sidewalk)
        
        if "SIDEWALK" in garden_name_full.upper():
            garden_data = garden_data[is_sidewalk]
        elif "STANDARD" in garden_name_full.upper():
            garden_data = garden_data[~is_sidewalk]
        elif "INFANT" in garden_name_full.upper():
            return "N/A"
        elif "MATTHEW" in garden_name_full.upper():
            pass # Use standard logic if no plot data
            
    # --- STANDARD SUBSECTION LOGIC ---
    elif sub_section:
        sub_mask_1 = garden_data[col_garden].astype(str).str.contains(sub_section, case=False, na=False)
        sub_mask_2 = pd.Series([False] * len(garden_data), index=garden_data.index)
        if col_section:
            sub_mask_2 = garden_data[col_section].astype(str).str.contains(sub_section, case=False, na=False)
        
        if (sub_mask_1 | sub_mask_2).any():
            garden_data = garden_data[sub_mask_1 | sub_mask_2]
    
    total = len(garden_data)
    if total == 0: return None
        
    avail_mask = garden_data[col_status].astype(str).str.contains('|'.join(status_avail), case=False, na=False)
    return (total - len(garden_data[avail_mask])) / total

def count_row_availability(df_inventory, garden_name, row_name, col_map):
    col_garden, col_row, col_status = col_map['Garden'], col_map['Row'], col_map['Status']
    if not col_garden or not col_row or not col_status: return None

    target_garden = super_clean_name(garden_name)
    if not target_garden: return None

    # Filter Garden
    inv_clean = df_inventory[col_garden].astype(str).apply(super_clean_name)
    garden_mask = inv_clean.str.contains(target_garden, case=False, na=False)
    garden_data = df_inventory[garden_mask]
    
    if garden_data.empty: return "N/A"

    # Filter Row (Exact Matching prevents "A" matching "AA")
    target_row = clean_row_name(row_name)
    if target_row == 'ALL': 
        row_data = garden_data
    else:
        # Use exact check on cleaned string to prevent over-matching
        row_mask = garden_data[col_row].astype(str).apply(lambda x: target_row == clean_row_name(x))
        row_data = garden_data[row_mask]
    
    if len(row_data) == 0: return None
    
    avail_mask = row_data[col_status].astype(str).str.contains('|'.join(status_avail), case=False, na=False)
    return len(row_data[avail_mask])

# --- 6. SURGICAL UPDATE ---
def surgical_update(inv_path, master_path, output_path):
    print("⚙️  Loading files...")
    
    try:
        header_row = find_inventory_header(inv_path)
        df_inv = pd.read_excel(inv_path, header=header_row)
        col_map = identify_columns(df_inv)
        print(f"   > Mapped Columns: {col_map}")
    except Exception as e: print(f"❌ Inventory Error: {e}"); return

    wb = load_workbook(master_path)
    
    # Styles
    HEADER_COLOR = "363636"
    HEADER_TEXT = "FFFFFF"
    ROW_ALT_COLOR = "F2F2F2"
    FONT_NAME = "Calibri"
    
    header_font = Font(name=FONT_NAME, size=11, bold=True, color=HEADER_TEXT)
    header_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
    alt_fill = PatternFill(start_color=ROW_ALT_COLOR, end_color=ROW_ALT_COLOR, fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(left=Side(style='thin', color="D9D9D9"), right=Side(style='thin', color="D9D9D9"), top=Side(style='thin', color="D9D9D9"), bottom=Side(style='thin', color="D9D9D9"))

    print("⚙️  Updating Sheets...")
    
    for ws in wb.worksheets:
        header_row = 1
        col_indices = {}
        row_col_idx = None  # To store the exact index of the "Row Name" column
        garden_col_idx = None # To store the exact index of the "Garden Name" column
        
        # 1. FIND HEADER & IDENTIFY KEY COLUMNS
        for r in range(1, 21):
            row_vals = [str(ws.cell(row=r, column=c).value).upper() for c in range(1, ws.max_column + 1)]
            if any(x in str(row_vals) for x in ['PRICE', 'GARDEN', 'LEVEL', 'ROW', 'SECTION', 'AVAIL']):
                header_row = r
                for c in range(1, ws.max_column + 1):
                    val = str(ws.cell(row=r, column=c).value).upper()
                    if val != 'NONE': 
                        col_indices[c] = val
                        # Identify specific columns to prevent overwrite bugs
                        if val in ['ROW', 'LEVEL', 'SECTION', 'CRYPT', 'NICHE', 'SPACE'] and not row_col_idx:
                            row_col_idx = c
                        if 'GARDEN' in val and not garden_col_idx:
                            garden_col_idx = c
                break
        
        sheet_name = ws.title
        search_name_spec = clean_sheet_name_specific(sheet_name)
        search_name_gen = clean_sheet_name_generic(sheet_name)
        
        if garden_exists_in_inventory(df_inv, search_name_spec, col_map):
            final_search_name = search_name_spec
        else:
            final_search_name = search_name_gen

        # 2. PROCESS ROWS
        data_row_count = 0
        
        for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row):
            r_idx = row[0].row
            row_name = None
            garden_name = None
            
            # Check secondary header
            is_secondary = False
            for cell in row:
                if cell.column in col_indices:
                    col_name = col_indices[cell.column]
                    if any(x in col_name for x in ['ROW', 'LEVEL', 'SECTION', 'PRODUCT']):
                        if str(cell.value).strip().upper() in ['ROW', 'LEVEL', 'SECTION', 'PRODUCT', 'STATION', 'NICHE #']:
                            is_secondary = True

            is_header = (r_idx == header_row) or is_secondary
            
            # --- GET KEYS RELIABLY ---
            if not is_header:
                # Use pre-identified columns if available, otherwise fallback to loop
                if row_col_idx:
                    row_name = str(ws.cell(row=r_idx, column=row_col_idx).value).strip()
                if garden_col_idx:
                    garden_name = str(ws.cell(row=r_idx, column=garden_col_idx).value).strip()
                
                # Fallback if specific columns weren't found (e.g. variable naming)
                if not row_name:
                    for cell in row:
                        if cell.column in col_indices:
                            name = col_indices[cell.column]
                            if any(x in name for x in ['ROW', 'LEVEL', 'SECTION', 'STATION', 'PRODUCT']):
                                row_name = str(cell.value).strip()
                if not garden_name:
                    for cell in row:
                        if cell.column in col_indices and 'GARDEN' in col_indices[cell.column]:
                            garden_name = str(cell.value).strip()

            # --- FORMATTING LOOP ---
            for cell in row:
                if isinstance(cell, MergedCell): continue

                col_name = col_indices.get(cell.column, "")
                cell.border = thin_border
                
                if is_header:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                else:
                    cell.font = Font(name=FONT_NAME, size=10, color="000000")
                    if data_row_count % 2 == 0: cell.fill = alt_fill
                    else: cell.fill = PatternFill(fill_type=None)
                    
                    is_desc = any(x in col_name for x in ['GARDEN', 'ROW', 'SECTION', 'LEVEL', 'PRODUCT', 'DESCRIPTION', 'LOCATION'])
                    cell.alignment = left_align if is_desc else center_align

                    # VALUE CLEANING (Triangle Killer)
                    val = cell.value
                    if val is not None:
                        # Strip %, $, commas
                        val_str = str(val).strip().replace('$', '').replace(',', '').replace('%', '')
                        
                        if any(x in col_name for x in ['PRICE', 'TOTAL', 'COST', 'PLAQUE']):
                            if val_str.replace('.', '', 1).isdigit():
                                try:
                                    cell.value = float(val_str)
                                    cell.number_format = '"$"#,##0'
                                    cell.data_type = 'n'
                                except: pass

                        elif any(x in col_name for x in ['%', 'SOLD', 'SCARCITY']) and 'QTY' not in col_name:
                            if val_str.replace('.', '', 1).isdigit():
                                try:
                                    cell.value = float(val_str)
                                    cell.number_format = '0%'
                                    cell.data_type = 'n'
                                except: pass
                        
                        elif any(x in col_name for x in ['QTY', 'AVAIL', 'STATUS']):
                            if str(cell.value).upper() == "SOLD OUT":
                                cell.font = Font(name=FONT_NAME, size=10, color="FF0000", bold=True)
                            elif val_str.replace('.', '', 1).isdigit():
                                try:
                                    cell.value = int(float(val_str))
                                    cell.number_format = '0'
                                    cell.data_type = 'n'
                                    if 0 < cell.value < 4:
                                        cell.font = Font(name=FONT_NAME, size=10, color="E26B0A", bold=True)
                                except: pass

            # --- DATA UPDATE ---
            if not is_header:
                if garden_name:
                    new_pct = calculate_percent_sold(df_inv, garden_name, col_map)
                    if new_pct is not None and new_pct != "N/A":
                        for c, name in col_indices.items():
                            if any(k in name for k in ['%', 'SOLD', 'SCARCITY']) and 'QTY' not in name: 
                                cell = ws.cell(row=r_idx, column=c)
                                if not isinstance(cell, MergedCell):
                                    cell.value = new_pct
                                    cell.number_format = '0%'
                                    cell.data_type = 'n'

                if row_name and row_name != 'None' and row_name != '':
                    count = count_row_availability(df_inv, final_search_name, row_name, col_map)
                    if count is not None and count != "N/A":
                        for c, name in col_indices.items():
                            if any(x in name for x in ['AVAIL', 'QTY', 'STATUS']):
                                cell = ws.cell(row=r_idx, column=c)
                                if not isinstance(cell, MergedCell):
                                    cell.alignment = center_align
                                    if count == 0:
                                        cell.value = "Sold Out"
                                        cell.font = Font(name=FONT_NAME, size=10, color="FF0000", bold=True)
                                    else:
                                        cell.value = int(count)
                                        cell.number_format = '0'
                                        cell.data_type = 'n'
                                        if count < 4:
                                            cell.font = Font(name=FONT_NAME, size=10, color="E26B0A", bold=True)
                
                data_row_count += 1

        # 3. AUTO-FIT
        for i, col in enumerate(ws.columns, 1):
            col_letter = get_column_letter(i)
            header_val = str(ws.cell(row=header_row, column=i).value).upper()
            
            if any(x in header_val for x in ['ROW', 'LEVEL', 'SECTION', 'PRODUCT']):
                ws.column_dimensions[col_letter].width = 12
                continue

            for cell in col:
                if cell.row >= header_row and cell.value and not isinstance(cell, MergedCell):
                    try: 
                        # Estimate width based on string length
                        val_len = len(str(cell.value))
                        curr_width = ws.column_dimensions[col_letter].width
                        # Approx check to avoid resetting custom widths too small
                        if val_len > 0:
                             ws.column_dimensions[col_letter].width = min(max(curr_width, (val_len + 2) * 1.1), 50)
                    except: pass

    wb.save(output_path)
    print(f"\nSUCCESS! 🚀\nFile saved: {output_path}")

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python3 update_v32.py [File1] [File2]")
    else:
        path_a = sys.argv[1].strip().replace("'", "").replace('"', "")
        path_b = sys.argv[2].strip().replace("'", "").replace('"', "")
        
        f1, f2 = os.path.basename(path_a).lower(), os.path.basename(path_b).lower()
        if 'inventory' in f1: inv_path, master_path = path_a, path_b
        elif 'inventory' in f2: inv_path, master_path = path_b, path_a
        else:
            print("❌ Error: One file must include 'Inventory' in the name.")
            sys.exit()

        folder = os.path.dirname(master_path)
        output_path = os.path.join(folder, 'Harpeth_Hills_Price_Book_FINAL.xlsx')
        surgical_update(inv_path, master_path, output_path)
