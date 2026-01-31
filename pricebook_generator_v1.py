"""
Harpeth Hills Dynamic Price Book Generator (v1.1)

Key goals
---------
- Bootstrap ONCE from Harpeth_Hills_Price_Book_REMADE.xlsx to create price_library.xlsx.
  * price_library.xlsx contains Baseline 2025 + BasePriceLocked (editable by you).
  * Missing Companion crypt prices are auto-filled on bootstrap using:
        CompanionCrypt = round_up_to_...995( 0.80 * (2 * SingleCrypt) )
    (You can override BasePriceLocked later.)

- Publish repeatedly:
  * Replace only the FaCTS export (Property - Property Inventory with Owner Details.xlsm)
  * Run publish -> outputs Harpeth_Hills_Price_Book_PUBLISHED.xlsx
  * Uses BasePriceLocked + tiered scarcity, availability from FaCTS.

Scarcity tiers (dynamic each publish)
-----------------------------------
- sold% >= 97%  -> +20%
- sold% >= 90%  -> +15%
- else          -> +0%

Notes
-----
- Sold-out rows ALWAYS print (pricing stays visible; availability shows 'Sold Out' in red).
- Building 7/8 covered/uncovered split is NOT derivable from FaCTS Space strings in the current export.
  This version uses row-theme buckets from FaCTS and prints both sections using the same scarcity.

Dependencies
------------
- pandas
- openpyxl

Usage
-----
python pricebook_generator.py bootstrap
python pricebook_generator.py publish
"""

from __future__ import annotations

import re
import math
from pathlib import Path
from typing import Dict, Tuple, List, Any

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# -------------------------------------------------
# FILES (fixed names so you only swap FaCTS each run)
# -------------------------------------------------
FACTS_XLSM = Path("Property - Property Inventory with Owner Details.xlsm")
MASTER_LISTING_XLSX = Path("Harpeth_Hills_Price_Book_REMADE.xlsx")
PRICE_LIBRARY_XLSX = Path("price_library.xlsx")
OUTPUT_XLSX = Path("Harpeth_Hills_Price_Book_PUBLISHED.xlsx")

# -----------------
# PRICING RULES
# -----------------
ROUND_TO = 995
DEFAULT_INCREASE_PCT = 0.05

TIER1_SOLD = 0.90
TIER1_UPLIFT = 0.15
TIER2_SOLD = 0.97
TIER2_UPLIFT = 0.20

COMPANION_DISCOUNT = 0.20  # 20% off two singles (default fill)


def round_up_to(x: float, base: int = ROUND_TO) -> int:
    """Round up to nearest multiple of base (995)."""
    if x <= 0:
        return 0
    return int(math.ceil(x / base) * base)


def round_up_end995(x: float) -> int:
    """Round up to next price ending in ...995."""
    return int(math.ceil((x + 5) / 1000.0) * 1000 - 5)


def scarcity_uplift(sold_pct: float) -> float:
    """Tiered scarcity uplift."""
    if sold_pct >= TIER2_SOLD:
        return TIER2_UPLIFT
    if sold_pct >= TIER1_SOLD:
        return TIER1_UPLIFT
    return 0.0


def final_price_from_base(base_price_locked: int, sold_pct: float) -> int:
    """Final = BasePriceLocked + tiered scarcity (no compounding)."""
    u = scarcity_uplift(sold_pct)
    if u <= 0:
        return int(base_price_locked)
    return round_up_to(base_price_locked * (1 + u), ROUND_TO)


# -----------------
# FORMATTING
# -----------------
DARK = "404040"
DARKER = "2B2B2B"
MID = "D9D9D9"
GRID = "BFBFBF"
RED = "C00000"

thin = Side(style="thin", color=GRID)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def set_widths(ws, widths: Dict[int, float]):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def title_bar(ws, title: str, last_col: int):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    c = ws.cell(1, 1, title)
    c.font = Font(bold=True, size=16, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26


def header_row(ws, headers: List[str], row: int = 2):
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row, col, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=DARKER)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    ws.row_dimensions[row].height = 18


def group_row(ws, row: int, text: str, last_col: int):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    c = ws.cell(row, 1, text)
    c.font = Font(bold=True)
    c.fill = PatternFill("solid", fgColor=MID)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = BORDER


def write_data_row(ws, row: int, values: List[Any], money_cols: set[int], avail_col: int):
    for col, v in enumerate(values, start=1):
        cell = ws.cell(row, col, v)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="left", vertical="center")

        if col in money_cols and isinstance(v, (int, float)):
            cell.number_format = '"$"#,##0'
            cell.alignment = Alignment(horizontal="right", vertical="center")

    # Availability formatting
    av = ws.cell(row, avail_col)
    av.alignment = Alignment(horizontal="center", vertical="center")
    if isinstance(av.value, (int, float)):
        av.number_format = "0"
    else:
        if str(av.value).strip().lower() == "sold out":
            av.font = Font(color=RED, bold=True)


# -----------------
# NORMALIZATION
# -----------------
DASH_CHARS = {
    "—": "–",
    "-": "–",
}


def norm_text(x: Any) -> str:
    if x is None:
        return ""
    s = str(x).strip()
    for k, v in DASH_CHARS.items():
        s = s.replace(k, v)
    # normalize repeated spaces
    s = re.sub(r"\s+", " ", s)
    return s


# -----------------
# FaCTS loader
# -----------------
REQUIRED_FACTS_COLS = {"Location", "Section", "Status", "Space", "Sales Item"}


def load_facts() -> pd.DataFrame:
    raw = pd.read_excel(FACTS_XLSM, sheet_name=0, engine="openpyxl")
    headers = [re.sub(r"\s+", " ", str(h)).strip() for h in raw.iloc[1].fillna("")]
    df = raw.iloc[2:].copy()
    df.columns = headers

    missing = REQUIRED_FACTS_COLS - set(df.columns)
    if missing:
        raise ValueError(f"FaCTS export is missing required columns: {', '.join(sorted(missing))}")

    df = df[df["Location"].notna()].copy()
    for c in ["Section", "Status", "Type", "Sales Item", "Space", "Right Types"]:
        if c in df.columns:
            df[c] = df[c].astype(str).str.strip()
    return df


def is_available(status: str) -> bool:
    return str(status).strip().lower() == "available"


# -----------------
# Sub-product availability parsers
# -----------------
# Mountain View crypts
ROW_THEME_MV = {"D": "D – Touch", "C": "C – Eye", "B": "B – Heart", "A": "A – Prayer"}
MV_CRYPT_RE = re.compile(
    r"Mountain View Mausoleum Crypts\s+(Upper Level|Lower Level)\s+Elevation\s+(\d+)\s+Level\s+([A-D])\s+Crypt\s+(\d+)",
    re.IGNORECASE,
)


def count_adjacent_pairs(nums: List[int]) -> int:
    nums = sorted(set(nums))
    used = set()
    pairs = 0
    for n in nums:
        if n in used:
            continue
        if (n + 1) in nums and (n + 1) not in used:
            used.add(n)
            used.add(n + 1)
            pairs += 1
    return pairs


def mv_buckets(facts: pd.DataFrame) -> Dict[Tuple[str, int, str, str], Dict[str, float]]:
    """Keyed by (level_band, elevation, row_theme, option) where option in {Single, Companion}."""
    mv = facts[facts["Section"].str.lower() == "mountain view"].copy()
    rows = []
    for _, r in mv.iterrows():
        m = MV_CRYPT_RE.search(r.get("Space", ""))
        if not m:
            continue
        band = m.group(1).title()
        elev = int(m.group(2))
        row_letter = m.group(3).upper()
        crypt_num = int(m.group(4))
        status = r.get("Status", "")
        sales_item = r.get("Sales Item", "")
        product = "Tandem" if "tandem" in sales_item.lower() else "Single"
        rows.append((band, elev, ROW_THEME_MV.get(row_letter, row_letter), crypt_num, product, status))

    if not rows:
        return {}

    inv = pd.DataFrame(rows, columns=["band", "elev", "theme", "crypt", "product", "status"])
    out: Dict[Tuple[str, int, str, str], Dict[str, float]] = {}

    for (band, elev, theme), g in inv[inv["product"] == "Single"].groupby(["band", "elev", "theme"]):
        total = len(g)
        avail = int(g["status"].map(is_available).sum())
        sold_pct = 1 - (avail / total) if total else 0.0
        out[(band, elev, theme, "Single")] = {"avail": avail, "total": total, "sold_pct": sold_pct}

        all_nums = g["crypt"].tolist()
        av_nums = g[g["status"].map(is_available)]["crypt"].tolist()
        total_pairs = count_adjacent_pairs(all_nums)
        avail_pairs = count_adjacent_pairs(av_nums)
        sold_pct_pairs = 1 - (avail_pairs / total_pairs) if total_pairs else 0.0
        out[(band, elev, theme, "Companion")] = {"avail": avail_pairs, "total": total_pairs, "sold_pct": sold_pct_pairs}

    return out


# Building 7/8/5 and Bell Tower
LS_CRYPT_RE = re.compile(
    r"(Last Supper Maus Bldg 7|Last Supper Maus Bldg 8|Last Supper Maus Bldg 5)\s+Crypt/Level\s+([0-9]+)([A-E])(?:-([0-9]+))?",
    re.IGNORECASE,
)
BT_CRYPT_RE = re.compile(
    r"(Bell Tower Mausoleum)\s+Crypt/Level\s+([0-9]+)([A-E])(?:-([0-9]+))?",
    re.IGNORECASE,
)

ROW_THEME_ABCDE = {
    "E": "E – Heavenly",
    "D": "D (Touch)",
    "C": "C (Eye)",
    "B": "B (Heart)",
    "A": "A (Prayer)",
}


def building_buckets_simple_by_row(
    facts: pd.DataFrame,
    facts_section_name: str,
    space_regex,
    include_tandem: bool = False,
) -> Dict[Tuple[str, str], Dict[str, float]]:
    """Buckets by row theme only (A–E), for Single and optionally Tandem."""
    sub = facts[facts["Section"] == facts_section_name].copy()
    rows = []
    for _, r in sub.iterrows():
        s = r.get("Space", "")
        m = space_regex.search(s)
        if not m:
            continue
        row_letter = m.group(3).upper()
        status = r.get("Status", "")
        sales_item = r.get("Sales Item", "")
        is_tandem = "tandem" in sales_item.lower()
        if is_tandem and not include_tandem:
            continue
        opt = "Tandem" if is_tandem else "Single"
        rows.append((ROW_THEME_ABCDE.get(row_letter, row_letter), opt, status))

    if not rows:
        return {}

    inv = pd.DataFrame(rows, columns=["theme", "option", "status"])
    out: Dict[Tuple[str, str], Dict[str, float]] = {}
    for (theme, opt), g in inv.groupby(["theme", "option"]):
        total = len(g)
        avail = int(g["status"].map(is_available).sum())
        sold_pct = 1 - (avail / total) if total else 0.0
        out[(theme, opt)] = {"avail": avail, "total": total, "sold_pct": sold_pct}
    return out


# -----------------
# PRICE LIBRARY bootstrap (includes companion fill)
# -----------------
MONEY_RE = re.compile(r"\$?\s*([0-9]{1,3}(?:,[0-9]{3})+|[0-9]+)")


def _to_int(v: Any) -> Optional[int]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        if isinstance(v, float) and math.isnan(v):
            return None
        return int(v)
    if isinstance(v, str):
        m = MONEY_RE.search(v)
        return int(m.group(1).replace(",", "")) if m else None
    return None


def bootstrap_price_library() -> None:
    if PRICE_LIBRARY_XLSX.exists():
        return

    wb = openpyxl.load_workbook(MASTER_LISTING_XLSX, data_only=True)

    skip = {
        "README",
        "Availability Dashboard",
        "Sold Out - Reference",
        "Sold Out \u2013 Reference",
        "Sold Out - Reference",
        "Needs Pricing Tables",
        "Cemetery Service Charges",
        "Pricing Issues Tracker",
    }

    def detect_price_sheets() -> List[str]:
        out = []
        for name in wb.sheetnames:
            if name in skip:
                continue
            ws = wb[name]
            found = False
            for r in range(1, 80):
                row_vals = [ws.cell(r, c).value for c in range(1, 10)]
                has_opt = any(isinstance(v, str) and v.strip().lower() == "option" for v in row_vals)
                has_crypt = any(isinstance(v, str) and v.strip().lower() == "crypt" for v in row_vals)
                if has_opt and has_crypt:
                    found = True
                    break
            if found:
                out.append(name)
        return out

    def parse_sheet(ws, product_family: str) -> List[Dict[str, Any]]:
        # locate header row
        header_row = None
        header = {}
        for r in range(1, 150):
            vals = [ws.cell(r, c).value for c in range(1, 12)]
            norm = [str(v).strip().lower() if isinstance(v, str) else "" for v in vals]
            if "option" in norm and "crypt" in norm:
                header_row = r
                for c, v in enumerate(norm, start=1):
                    if v:
                        header[v] = c
                break
        if not header_row:
            return []

        opt_c = header.get("option")
        crypt_c = header.get("crypt")
        label_c = header.get("row") or header.get("section") or header.get("product") or header.get("garden")
        front_c = header.get("crypt front") or header.get("plaque(s)") or header.get("niche front") or header.get("laser etching")
        total_c = header.get("total") or header.get("total price") or header.get("all-in total")
        avail_c = header.get("availability") or header.get("# available") or header.get("available")

        # detect plaque prices in header text
        single_plaque = None
        tandem_plaque = None
        companion_plaque = None
        for r in range(1, 35):
            for c in range(1, 10):
                v = ws.cell(r, c).value
                if not isinstance(v, str):
                    continue
                s = v.lower()
                if "single" in s and "plaque" in s:
                    m = MONEY_RE.search(v)
                    if m:
                        single_plaque = int(m.group(1).replace(",", ""))
                if "tandem" in s and "plaque" in s:
                    m = MONEY_RE.search(v)
                    if m:
                        tandem_plaque = int(m.group(1).replace(",", ""))
                if "companion" in s and "plaque" in s:
                    m = MONEY_RE.search(v)
                    if m:
                        companion_plaque = int(m.group(1).replace(",", ""))

        group = None
        theme = None
        recs: List[Dict[str, Any]] = []

        for r in range(header_row + 1, ws.max_row + 1):
            label = ws.cell(r, label_c).value if label_c else None
            opt = ws.cell(r, opt_c).value if opt_c else None
            crypt = ws.cell(r, crypt_c).value if crypt_c else None

            if label is None and opt is None and crypt is None:
                continue

            if isinstance(label, str):
                s = norm_text(label)
                if s.lower().startswith("elevation") or s.upper() in ["COVERED", "UNCOVERED", "ALL LEVELS"] or "COVERED" in s.upper() or "UNCOVERED" in s.upper():
                    group = s
                    theme = None
                    continue
                # treat these as "theme" rows
                if "–" in s or ("(" in s and any(x in s for x in ["Touch", "Eye", "Heart", "Prayer", "Heavenly"])):
                    theme = s

            if not isinstance(opt, str):
                continue
            opt_s = norm_text(opt)
            if opt_s not in ["Single", "Companion", "Tandem"]:
                continue

            recs.append(
                {
                    "product_family": product_family,
                    "group": group,
                    "theme": theme,
                    "option": opt_s,
                    "baseline_2025_crypt": _to_int(crypt),
                    "baseline_2025_front": _to_int(ws.cell(r, front_c).value) if front_c else None,
                    "baseline_2025_total": _to_int(ws.cell(r, total_c).value) if total_c else None,
                    "availability_text": ws.cell(r, avail_c).value if avail_c else None,
                    "single_plaque": single_plaque,
                    "tandem_plaque": tandem_plaque,
                    "companion_plaque": companion_plaque,
                }
            )

        return recs

    price_sheets = detect_price_sheets()
    all_rows: List[Dict[str, Any]] = []
    for name in price_sheets:
        all_rows.extend(parse_sheet(wb[name], name))

    df = pd.DataFrame(all_rows)
    if df.empty:
        raise RuntimeError("No price-table sheets detected in master listing workbook.")

    # normalize theme/group text to avoid mismatch later
    df["group"] = df["group"].apply(norm_text)
    df["theme"] = df["theme"].apply(norm_text)
    df["option"] = df["option"].apply(norm_text)

    # Fill missing companion crypt prices from Single crypts: 0.8 * (2 * single), rounded to ...995
    single_map = {
        (r.product_family, r.group, r.theme): r.baseline_2025_crypt
        for r in df[df["option"] == "Single"].itertuples()
        if pd.notna(r.baseline_2025_crypt)
    }
    single_front_map = {
        (r.product_family, r.group, r.theme): r.baseline_2025_front
        for r in df[df["option"] == "Single"].itertuples()
        if pd.notna(r.baseline_2025_front)
    }

    for idx, r in df.iterrows():
        if r["option"] != "Companion":
            continue
        key = (r["product_family"], r["group"], r["theme"])
        single_crypt = single_map.get(key)
        if pd.isna(r["baseline_2025_crypt"]) and single_crypt is not None:
            df.at[idx, "baseline_2025_crypt"] = round_up_end995(2 * single_crypt * (1 - COMPANION_DISCOUNT))

        if pd.isna(r["baseline_2025_front"]):
            if pd.notna(r.get("companion_plaque")):
                df.at[idx, "baseline_2025_front"] = int(r["companion_plaque"])
            elif pd.notna(r.get("single_plaque")):
                df.at[idx, "baseline_2025_front"] = int(r["single_plaque"]) * 2
            else:
                sf = single_front_map.get(key)
                if sf is not None and not pd.isna(sf):
                    df.at[idx, "baseline_2025_front"] = int(sf) * 2

    # Fill totals if missing
    mask = df["baseline_2025_total"].isna() & df["baseline_2025_crypt"].notna() & df["baseline_2025_front"].notna()
    df.loc[mask, "baseline_2025_total"] = df.loc[mask, "baseline_2025_crypt"] + df.loc[mask, "baseline_2025_front"]

    # Initialize BasePriceLocked = baseline * 1.05, rounded up to 995
    df["increase_pct"] = DEFAULT_INCREASE_PCT
    for part in ["crypt", "front", "total"]:
        src = f"baseline_2025_{part}"
        dst = f"base_price_locked_{part}"
        df[dst] = pd.to_numeric(df[src], errors="coerce").apply(
            lambda x: round_up_to(x * (1 + DEFAULT_INCREASE_PCT), ROUND_TO) if pd.notna(x) else None
        )

    policy = pd.DataFrame(
        [
            {
                "round_to": ROUND_TO,
                "default_increase_pct": DEFAULT_INCREASE_PCT,
                "companion_discount_pct": COMPANION_DISCOUNT,
                "tier1_sold_pct": TIER1_SOLD,
                "tier1_uplift": TIER1_UPLIFT,
                "tier2_sold_pct": TIER2_SOLD,
                "tier2_uplift": TIER2_UPLIFT,
                "note": "Edit BasePriceLocked_* anytime. Publish runs only update scarcity + availability.",
            }
        ]
    )

    with pd.ExcelWriter(PRICE_LIBRARY_XLSX, engine="openpyxl") as w:
        policy.to_excel(w, sheet_name="Pricing_Policy", index=False)
        df.to_excel(w, sheet_name="Price_Library", index=False)

    print("Created price_library.xlsx")


# -----------------
# PUBLISH
# -----------------

def publish() -> None:
    if not PRICE_LIBRARY_XLSX.exists():
        bootstrap_price_library()

    facts = load_facts()
    lib = pd.read_excel(PRICE_LIBRARY_XLSX, sheet_name="Price_Library")

    # normalize key text fields
    for col in ["product_family", "group", "theme", "option"]:
        if col in lib.columns:
            lib[col] = lib[col].apply(norm_text)

    # Compute scarcity buckets from FaCTS
    mv = mv_buckets(facts)
    b7 = building_buckets_simple_by_row(facts, "Last Supper Maus Bldg 7", LS_CRYPT_RE, include_tandem=False)
    b8 = building_buckets_simple_by_row(facts, "Last Supper Maus Bldg 8", LS_CRYPT_RE, include_tandem=False)
    bt = building_buckets_simple_by_row(facts, "Bell Tower Mausoleum", BT_CRYPT_RE, include_tandem=True)

    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)

    def build_sheet(product_family: str, title: str, headers: List[str], widths: Dict[int, float]):
        ws = out_wb.create_sheet(product_family[:31])
        title_bar(ws, title, last_col=len(headers))
        header_row(ws, headers, row=2)
        ws.freeze_panes = "A3"
        set_widths(ws, widths)

        df = lib[lib["product_family"] == norm_text(product_family)].copy()
        r = 3

        if product_family.startswith("Mountain View"):
            elevs = sorted(
                [e for e in df["group"].dropna().unique() if str(e).lower().startswith("elevation")],
                key=lambda x: int(re.search(r"(\d+)", str(x)).group(1)),
            )
            theme_order = ["D – Touch", "C – Eye", "B – Heart", "A – Prayer"]

            for g in elevs:
                group_row(ws, r, str(g), last_col=len(headers))
                r += 1
                elev = int(re.search(r"(\d+)", str(g)).group(1))
                band = "Upper Level" if "Upper" in product_family else "Lower Level"

                for theme in theme_order:
                    for opt in ["Single", "Companion"]:
                        row_match = df[(df["group"] == norm_text(g)) & (df["theme"] == norm_text(theme)) & (df["option"] == opt)]
                        if len(row_match) != 1:
                            continue
                        rowi = row_match.iloc[0]
                        bucket = mv.get((band, elev, theme, opt), {"avail": 0, "total": 0, "sold_pct": 0.0})
                        sold_pct = float(bucket["sold_pct"])
                        avail = int(bucket["avail"])
                        availability = avail if avail > 0 else "Sold Out"

                        crypt = final_price_from_base(int(rowi["base_price_locked_crypt"]), sold_pct) if pd.notna(rowi["base_price_locked_crypt"]) else None
                        front = final_price_from_base(int(rowi["base_price_locked_front"]), sold_pct) if pd.notna(rowi["base_price_locked_front"]) else None
                        total = (crypt + front) if (crypt is not None and front is not None) else None

                        theme_label = theme if opt == "Single" else None
                        write_data_row(ws, r, [theme_label, opt, crypt, front, total, availability], money_cols={3, 4, 5}, avail_col=6)
                        r += 1

                r += 1

        else:
            # grouped sections (COVERED / UNCOVERED) from library
            groups = [g for g in df["group"].dropna().unique()]

            def grp_sort(x):
                s = str(x).upper()
                if "UNCOVERED" in s:
                    return (0, s)
                if "COVERED" in s:
                    return (1, s)
                return (2, s)

            groups = sorted(groups, key=grp_sort)

            # Theme order is whatever exists in this product family; sorted by A-E then others
            theme_rank = {
                "E": 0,
                "D": 1,
                "C": 2,
                "B": 3,
                "A": 4,
            }

            themes = [t for t in df["theme"].dropna().unique()]

            def theme_key(t):
                tt = str(t)
                m = re.match(r"^([A-E])\b", tt)
                if m:
                    return (0, theme_rank.get(m.group(1), 9), tt)
                return (1, 9, tt)

            themes = sorted(themes, key=theme_key)

            if product_family == "Building 7 Mausoleum":
                bucket_map = b7
            elif product_family == "Building 8 Mausoleum":
                bucket_map = b8
            elif product_family == "Bell Tower Mausoleum":
                bucket_map = bt
            else:
                bucket_map = {}

            for g in groups:
                group_row(ws, r, str(g), last_col=len(headers))
                r += 1

                for theme in themes:
                    for opt in ["Single", "Tandem", "Companion"]:
                        row_match = df[(df["group"] == norm_text(g)) & (df["theme"] == norm_text(theme)) & (df["option"] == opt)]
                        if len(row_match) != 1:
                            continue
                        rowi = row_match.iloc[0]

                        if opt == "Companion":
                            # IMPORTANT FIX:
                            # We do NOT reuse Single availability count as Companion availability.
                            # Until we have companion-unit mapping here, we always display pricing and show Sold Out.
                            b = bucket_map.get((norm_text(theme), "Single"), {"sold_pct": 0.0})
                            sold_pct = float(b.get("sold_pct", 0.0))
                            availability = "Sold Out"
                        else:
                            b = bucket_map.get((norm_text(theme), opt), {"avail": 0, "sold_pct": 0.0})
                            sold_pct = float(b.get("sold_pct", 0.0))
                            avail = int(b.get("avail", 0))
                            availability = avail if avail > 0 else "Sold Out"

                        crypt = final_price_from_base(int(rowi["base_price_locked_crypt"]), sold_pct) if pd.notna(rowi["base_price_locked_crypt"]) else None
                        front = final_price_from_base(int(rowi["base_price_locked_front"]), sold_pct) if pd.notna(rowi["base_price_locked_front"]) else None
                        total = (crypt + front) if (crypt is not None and front is not None) else None

                        theme_label = theme if opt == "Single" else None
                        write_data_row(ws, r, [theme_label, opt, crypt, front, total, availability], money_cols={3, 4, 5}, avail_col=6)
                        r += 1

                r += 1

        return ws

    # Build sheets (extend as you add more product families to the library)
    build_sheet(
        "Mountain View - Upper Level",
        "MOUNTAIN VIEW MAUSOLEUM — UPPER LEVEL",
        ["Row", "Option", "Crypt", "Crypt Front", "Total", "Availability"],
        {1: 22, 2: 12, 3: 14, 4: 14, 5: 14, 6: 12},
    )
    build_sheet(
        "Mountain View - Lower Level",
        "MOUNTAIN VIEW MAUSOLEUM — LOWER LEVEL",
        ["Row", "Option", "Crypt", "Crypt Front", "Total", "Availability"],
        {1: 22, 2: 12, 3: 14, 4: 14, 5: 14, 6: 12},
    )
    build_sheet(
        "Building 7 Mausoleum",
        "BUILDING 7 MAUSOLEUM",
        ["Row", "Option", "Crypt", "Crypt Front", "Total", "Availability"],
        {1: 22, 2: 12, 3: 14, 4: 14, 5: 14, 6: 12},
    )
    build_sheet(
        "Building 8 Mausoleum",
        "BUILDING 8 MAUSOLEUM",
        ["Section", "Option", "Crypt", "Crypt Front", "Total", "Availability"],
        {1: 28, 2: 12, 3: 14, 4: 14, 5: 14, 6: 12},
    )
    build_sheet(
        "Bell Tower Mausoleum",
        "BELL TOWER MAUSOLEUM",
        ["Row", "Option", "Crypt", "Crypt Front", "Total", "Availability"],
        {1: 22, 2: 12, 3: 14, 4: 14, 5: 14, 6: 12},
    )

    out_wb.save(OUTPUT_XLSX)
    print(f"Wrote: {OUTPUT_XLSX}")


# -----------------
# CLI
# -----------------
if __name__ == "__main__":
    import sys

    cmd = sys.argv[1].lower() if len(sys.argv) > 1 else "publish"
    if cmd == "bootstrap":
        bootstrap_price_library()
        print("Bootstrap complete. Edit BasePriceLocked in price_library.xlsx as needed.")
    elif cmd == "publish":
        publish()
    else:
        raise SystemExit("Usage: python pricebook_generator.py bootstrap|publish")
