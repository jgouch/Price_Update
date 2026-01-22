import pandas as pd
import re
import os
import sys
import warnings
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- 1. SETUP ---
warnings.simplefilter(action='ignore', category=FutureWarning)

# --- 2. LOGIC & CLEANING FUNCTIONS ---

def clean_sheet_name_specific(sheet_name):
    """Minimal cleaning. Keeps 'Columbarium' etc."""
    name = re.sub(r'^\d+_', '', sheet_name)
    remove_words = ['Mausoleum', 'Bldg', 'Building']
    for word in remove_words:
        name = name.replace(word, '')
    return name.strip()

def clean_sheet_name_generic(sheet_name):
    """Aggressive cleaning to find base garden name."""
    name = clean_sheet_name_specific(sheet_name)
    remove_words = ['Columbarium', 'Niches', 'Garden']
    for word in remove_words:
        name = name.replace(word, '')
    return name.strip()

def clean_row_name(row_str):
    s = str(row_str).strip().upper()
    if 'ALL LEVEL' in s: return 'ALL'
    s = re.sub(r'\(.*?\)', '', s)
    s = s.replace('UNCOVERED', '').replace('COVERED', '').replace('ELEVATION', '')
    s = s.replace('—', '-').replace('–', '-')
    if '-' in s: s = s.split('-')[0]
    return s.strip()

def identify_columns(df):
    cols = [str(c) for c in df.columns]
    mapping = {'Garden': None, 'Row': None, 'Status': None}
    
    for c in cols:
        if 'GARDEN' in c.upper() or 'GROUP' in c.upper() or 'LOCATION' in c.upper():
            mapping['Garden'] = c
            break

    candidates = [c for c in cols if any(x in c.upper() for x in ['SECTION', 'ROW', 'BLOCK', 'LOT', 'TIER'])]
    if candidates:
        best = next((x for x in candidates if 'SECTION' in x.upper()), None)
        if not best: best = next((x for x in candidates if 'ROW' in x.upper()), None)
        mapping['Row'] = best if best else candidates[0]

    for c in cols:
        if 'STATUS' in c.upper() or 'STATE' in c.upper():
            mapping['Status'] = c
            break
            
    return mapping

def garden_exists_in_inventory(df_inventory, garden_name, col_map):
    col_garden = col_map['Garden']
    mask = df_inventory[col_garden].astype(str).str.contains(garden_name, case=False, na=False)
    return mask.any()

def calculate_percent_sold(df_inventory, garden_name_full, col_map):
    col_garden = col_map['Garden']
    col_section = col_map['Row']
    col_status = col_map['Status']
    status_avail = ['Available', 'Serviceable', 'For Sale', 'Vacant']
    
    if '–' in garden_name_full: parts = garden_name_full.split('–')
    elif '-' in garden_name_full: parts = garden_name_full.split('-')
    else: parts = [garden_name_full]
    
    main_garden = parts[0].strip()
    sub_section = parts[1].strip() if len(parts) > 1 else None

    garden_mask = df_inventory[col_garden].astype(str).str.contains(main_garden, case=False, na=False)
    garden_data = df_inventory[garden_mask]
    
    if sub_section and not garden_data.empty:
        section_mask = garden_data[col_section].astype(str).str.contains(sub_section, case=False, na=False)
        if section_mask.any():
            garden_data = garden_data[section_mask]
    
    total = len(garden_data)
    if total == 0: return None
        
    avail_mask = garden_data[col_status].astype(str).str.contains('|'.join(status_avail), case=False, na=False)
    avail_count = len(garden_data[avail_mask])
    
    return (total - avail_count) / total

def count_row_availability(df_inventory, garden_name, row_name, col_map):
    col_garden = col_map['Garden']
    col_row = col_map['Row']
    col_status = col_map['Status']
    status_avail = ['Available', 'Serviceable', 'For Sale', 'Vacant']

    garden_mask = df_inventory[col_garden].astype(str).str.contains(garden_name, case=False, na=False)
    garden_data = df_inventory[garden_mask]
    
    if garden_data.empty: return "N/A"

    target = clean_row_name(row_name)
    if target == 'ALL':
        row_data = garden_data
    else:
        row_mask = garden_data[col_row].astype(str).apply(clean_row_name) == target
        row_data = garden_data[row_mask]
    
    if len(row_data) == 0: return None
    
    avail_mask = row_data[col_status].astype(str).str.contains('|'.join(status_avail), case=False, na=False)
    return len(row_data[avail_mask])

# --- 3. FORMATTING FUNCTION ---

def apply_professional_formatting(file_path):
    print("\n🎨 Applying Professional Formatting...")
    
    # Colors
    HEADER_COLOR = "363636"  # Dark Grey
    HEADER_TEXT = "FFFFFF"   # White
    ROW_ALT_COLOR = "F2F2F2" # Light Grey
    FONT_NAME = "Calibri"

    wb = load_workbook(file_path)
    
    header_font = Font(name=FONT_NAME, size=11, bold=True, color=HEADER_TEXT)
    header_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
    thin_border = Border(left=Side(style='thin', color="D9D9D9"), right=Side(style='thin', color="D9D9D9"), top=Side(style='thin', color="D9D9D9"), bottom=Side(style='thin', color="D9D9D9"))
    align_center = Alignment(horizontal="center", vertical="center")
    
    for ws in wb.worksheets:
        # Auto-fit Columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[column].width = min((max_length + 2) * 1.1, 50)

        # Row Styling
        for i, row in enumerate(ws.iter_rows()):
            for cell in row:
                cell.font = Font(name=FONT_NAME, size=10)
                cell.border = thin_border
                
                # Header
                if i == 0:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = align_center
                # Data Rows
                else:
                    if i % 2 == 0:
                        cell.fill = PatternFill(start_color=ROW_ALT_COLOR, end_color=ROW_ALT_COLOR, fill_type="solid")
                    
                    # Number Formats
                    header_val = ws.cell(row=1, column=cell.column).value
                    if header_val:
                        h = str(header_val).upper()
                        if 'PRICE' in h or 'TOTAL' in h or 'RIGHT' in h:
                            cell.number_format = '"$"#,##0'
                            cell.alignment = align_center
                        elif '%' in h or 'SOLD' in h:
                            cell.number_format = '0%'
                            cell.alignment = align_center
                        elif 'QTY' in h or 'AVAIL' in h:
                            cell.alignment = align_center
                            if str(cell.value).lower() == "sold out":
                                cell.font = Font(name=FONT_NAME, size=10, color="FF0000", bold=True)
                            elif isinstance(cell.value, (int, float)) and cell.value > 0:
                                cell.font = Font(name=FONT_NAME, size=10, color="000000", bold=True)

    wb.save(file_path)
    print("✨ Formatting Complete.")

# --- 4. MAIN EXECUTION ---

def main():
    if len(sys.argv) < 3:
        print("Usage: python3 update_v11.py [File1] [File2]")
        return

    path_a = sys.argv[1].strip().replace("'", "").replace('"', "")
    path_b = sys.argv[2].strip().replace("'", "").replace('"', "")

    # Auto-Detect Files
    f1 = os.path.basename(path_a).lower()
    f2 = os.path.basename(path_b).lower()
    
    if 'inventory' in f1 and 'inventory' not in f2:
        inv_path, master_path = path_a, path_b
    elif 'inventory' in f2 and 'inventory' not in f1:
        inv_path, master_path = path_b, path_a
    else:
        print("❌ Error: Could not verify files. One filename must include 'Inventory'.")
        return

    print(f"\n📂 Inventory: {os.path.basename(inv_path)}")
    print(f"📘 Master Book: {os.path.basename(master_path)}")

    # 1. READ INVENTORY
    try:
        df_inv = pd.read_excel(inv_path, header=2)
        col_map = identify_columns(df_inv)
        print(f"✅ Inventory Loaded & Mapped")
        if None in col_map.values():
            print("❌ Error: Missing columns in inventory.")
            return
    except Exception as e:
        print(f"❌ Error reading inventory: {e}")
        return

    # 2. READ MASTER & UPDATE
    try:
        sheet_dict = pd.read_excel(master_path, sheet_name=None)
    except Exception as e:
        print(f"❌ Error reading master book: {e}")
        return

    print("⚙️  Processing Updates...")
    folder = os.path.dirname(master_path)
    output_path = os.path.join(folder, 'Harpeth_Hills_Price_Book_FINAL.xlsx')

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for sheet_name, df in sheet_dict.items():
            
            # Name Matching Logic
            search_name_specific = clean_sheet_name_specific(sheet_name)
            search_name_generic = clean_sheet_name_generic(sheet_name)
            
            if garden_exists_in_inventory(df_inv, search_name_specific, col_map):
                final_search_name = search_name_specific
            else:
                final_search_name = search_name_generic

            cols = [str(c).upper() for c in df.columns]
            
            # Update % Sold
            if any('%' in c for c in cols) and 'GARDEN' in cols:
                garden_col = next(c for c in df.columns if str(c).upper() == 'GARDEN')
                sold_col = next(c for c in df.columns if '%' in str(c))
                for idx, row in df.iterrows():
                    new_pct = calculate_percent_sold(df_inv, str(row[garden_col]), col_map)
                    if new_pct is not None: df.at[idx, sold_col] = new_pct

            # Update Counts
            row_candidates = [c for c in df.columns if any(x in str(c).upper() for x in ['ROW', 'LEVEL', 'SECTION', 'STATION', 'PRODUCT'])]
            qty_candidates = [c for c in df.columns if any(x in str(c).upper() for x in ['AVAIL', 'QTY', 'STATUS'])]
            
            if row_candidates and qty_candidates:
                row_col = row_candidates[0]
                qty_col = qty_candidates[0]
                for idx, row in df.iterrows():
                    row_val = row[row_col]
                    if pd.isna(row_val) or str(row_val).strip() == '': continue
                    count = count_row_availability(df_inv, final_search_name, str(row_val), col_map)
                    if count is not None and count != "N/A":
                        df.at[idx, qty_col] = "Sold Out" if count == 0 else count

            df.to_excel(writer, sheet_name=sheet_name, index=False)

    # 3. APPLY FORMATTING
    apply_professional_formatting(output_path)

    print("-" * 50)
    print(f"SUCCESS! 🚀\nFile saved: {output_path}")
    print("-" * 50)

if __name__ == "__main__":
    main()
