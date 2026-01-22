import pandas as pd
import re
import os
import sys
import warnings
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# --- 1. SETUP ---
warnings.simplefilter(action='ignore', category=FutureWarning)

# --- 2. CLEANING FUNCTIONS ---
def clean_sheet_name_specific(sheet_name):
    name = re.sub(r'^\d+_', '', sheet_name)
    for word in ['Mausoleum', 'Bldg', 'Building']:
        name = name.replace(word, '')
    return name.strip()

def clean_sheet_name_generic(sheet_name):
    name = clean_sheet_name_specific(sheet_name)
    for word in ['Columbarium', 'Niches', 'Garden']:
        name = name.replace(word, '')
    return name.strip()

def clean_row_name(row_str):
    s = str(row_str).strip().upper()
    if 'ALL LEVEL' in s: return 'ALL'
    s = re.sub(r'\(.*?\)', '', s)
    s = s.replace('UNCOVERED', '').replace('COVERED', '').replace('ELEVATION', '')
    s = s.replace('—', '-').replace('–', '-')
    if '-' in s: s = s.split('-')[0]
    return s.strip()

# --- 3. MAPPING & SCANNING ---
def identify_columns(df):
    cols = [str(c) for c in df.columns]
    mapping = {'Garden': None, 'Row': None, 'Status': None}
    
    for c in cols:
        if 'GARDEN' in c.upper() or 'GROUP' in c.upper() or 'LOCATION' in c.upper():
            mapping['Garden'] = c; break
    
    candidates = [c for c in cols if any(x in c.upper() for x in ['SECTION', 'ROW', 'BLOCK', 'LOT', 'TIER'])]
    if candidates:
        best = next((x for x in candidates if 'SECTION' in x.upper()), None)
        mapping['Row'] = best if best else candidates[0]

    for c in cols:
        if 'STATUS' in c.upper() or 'STATE' in c.upper():
            mapping['Status'] = c; break
            
    return mapping

def garden_exists_in_inventory(df_inventory, garden_name, col_map):
    col_garden = col_map['Garden']
    if not col_garden:
        return False
    mask = df_inventory[col_garden].astype(str).str.contains(garden_name, case=False, na=False)
    return mask.any()

# --- 4. CALCULATIONS ---
def calculate_percent_sold(df_inventory, garden_name_full, col_map):
    col_garden, col_section, col_status = col_map['Garden'], col_map['Row'], col_map['Status']
    if not col_garden or not col_status:
        return None
    status_avail = ['Available', 'Serviceable', 'For Sale', 'Vacant']
    status_pattern = '|'.join(re.escape(status) for status in status_avail)
    
    parts = re.split(r'[-–]', garden_name_full)
    main_garden = parts[0].strip()
    sub_section = parts[1].strip() if len(parts) > 1 else None

    garden_mask = df_inventory[col_garden].astype(str).str.contains(main_garden, case=False, na=False)
    garden_data = df_inventory[garden_mask]
    
    if sub_section and not garden_data.empty and col_section:
        section_mask = garden_data[col_section].astype(str).str.contains(sub_section, case=False, na=False)
        if section_mask.any(): garden_data = garden_data[section_mask]
    
    total = len(garden_data)
    if total == 0: return None
        
    avail_mask = garden_data[col_status].astype(str).str.contains(status_pattern, case=False, na=False)
    return (total - len(garden_data[avail_mask])) / total

def count_row_availability(df_inventory, garden_name, row_name, col_map):
    col_garden, col_row, col_status = col_map['Garden'], col_map['Row'], col_map['Status']
    if not col_garden or not col_row or not col_status:
        return None
    status_avail = ['Available', 'Serviceable', 'For Sale', 'Vacant']
    status_pattern = '|'.join(re.escape(status) for status in status_avail)

    garden_mask = df_inventory[col_garden].astype(str).str.contains(garden_name, case=False, na=False)
    garden_data = df_inventory[garden_mask]
    
    if garden_data.empty: return "N/A"

    target = clean_row_name(row_name)
    if target == 'ALL': row_data = garden_data
    else:
        row_mask = garden_data[col_row].astype(str).apply(clean_row_name) == target
        row_data = garden_data[row_mask]
    
    if len(row_data) == 0: return None
    
    avail_mask = row_data[col_status].astype(str).str.contains(status_pattern, case=False, na=False)
    return len(row_data[avail_mask])

# --- 5. SURGICAL UPDATE (Preserves Titles) ---
def surgical_update(inv_path, master_path, output_path):
    print("⚙️  Starting Surgical Update (Handling Multiple Headers)...")
    
    # A. Load Inventory
    try:
        df_inv = pd.read_excel(inv_path, header=2)
        col_map = identify_columns(df_inv)
    except Exception as e: print(f"❌ Inventory Error: {e}"); return

    # B. Load Master Workbook (OpenPyXL)
    wb = load_workbook(master_path)
    
    # Styles
    HEADER_COLOR = "363636" # Dark Grey
    HEADER_TEXT = "FFFFFF"  # White
    ROW_ALT_COLOR = "F2F2F2" # Light Grey
    FONT_NAME = "Calibri"
    
    header_font = Font(name=FONT_NAME, size=11, bold=True, color=HEADER_TEXT)
    header_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
    alt_fill = PatternFill(start_color=ROW_ALT_COLOR, end_color=ROW_ALT_COLOR, fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin', color="D9D9D9"), right=Side(style='thin', color="D9D9D9"), top=Side(style='thin', color="D9D9D9"), bottom=Side(style='thin', color="D9D9D9"))

    for ws in wb.worksheets:
        # 1. FIND FIRST HEADER ROW
        header_row = 1
        col_indices = {}
        
        for r in range(1, 21):
            row_vals = [str(ws.cell(row=r, column=c).value).upper() for c in range(1, ws.max_column + 1)]
            if any(x in cell_val for cell_val in row_vals for x in ['PRICE', 'GARDEN', 'LEVEL', 'ROW', 'SECTION', 'AVAIL']):
                header_row = r
                for c in range(1, ws.max_column + 1):
                    val = str(ws.cell(row=r, column=c).value).upper()
                    if val != 'NONE': col_indices[c] = val
                break
        
        # Determine Search Name
        sheet_name = ws.title
        search_name_spec = clean_sheet_name_specific(sheet_name)
        search_name_gen = clean_sheet_name_generic(sheet_name)
        final_search_name = search_name_spec if garden_exists_in_inventory(df_inv, search_name_spec, col_map) else search_name_gen

        # 2. ITERATE ROWS
        data_row_count = 0
        
        for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row):
            # Identify Key Columns
            r_idx = row[0].row
            row_name = None
            garden_name = None
            is_secondary_header = False
            
            # --- PRE-SCAN: Check if this is a Secondary Header Row ---
            # (e.g., if Row Column says "Level" or "Section" again)
            for cell in row:
                if cell.column in col_indices:
                    col_name = col_indices[cell.column]
                    if any(x in col_name for x in ['ROW', 'LEVEL', 'SECTION', 'PRODUCT']):
                        if str(cell.value).strip().upper() in ['ROW', 'LEVEL', 'SECTION', 'PRODUCT', 'STATION', 'NICHE #']:
                            is_secondary_header = True

            is_header = (row[0].row == header_row) or is_secondary_header
            
            # Loop Cells
            for cell in row:
                col_name = col_indices.get(cell.column, "")
                
                # FORMATTING
                cell.border = thin_border
                cell.font = Font(name=FONT_NAME, size=10)
                
                if is_header:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                else:
                    # Zebra Stripe
                    if data_row_count % 2 == 0:
                        cell.fill = alt_fill
                    else:
                        cell.fill = PatternFill(fill_type=None)
                    
                    # Number Formatting
                    if any(x in col_name for x in ['PRICE', 'TOTAL', 'RIGHT', 'COST', 'PLAQUE', 'ETCHING']):
                        cell.number_format = '"$"#,##0'
                        cell.alignment = center_align
                        if cell.value and str(cell.value).replace('.','',1).isdigit():
                            cell.value = float(str(cell.value))
                            
                    elif any(x in col_name for x in ['%', 'SOLD']) and 'QTY' not in col_name:
                        cell.number_format = '0%'
                        cell.alignment = center_align
                        
                    elif any(x in col_name for x in ['QTY', 'AVAIL', 'STATUS']):
                        cell.alignment = center_align
                        if str(cell.value).upper() == "SOLD OUT":
                            cell.font = Font(name=FONT_NAME, size=10, color="FF0000", bold=True)
                
                # DATA CAPTURE
                if not is_header:
                    val_str = str(cell.value).strip()
                    if any(x in col_name for x in ['ROW', 'LEVEL', 'SECTION', 'STATION', 'PRODUCT']):
                        row_name = val_str
                    if 'GARDEN' in col_name:
                        garden_name = val_str
            
            # UPDATE LOGIC
            if not is_header:
                if garden_name:
                    new_pct = calculate_percent_sold(df_inv, garden_name, col_map)
                    if new_pct is not None:
                        for c, name in col_indices.items():
                            if '%' in name: ws.cell(row=r_idx, column=c).value = new_pct

                if row_name and row_name.strip().lower() != 'none':
                    count = count_row_availability(df_inv, final_search_name, row_name, col_map)
                    if count is not None and count != "N/A":
                        for c, name in col_indices.items():
                            if any(x in name for x in ['AVAIL', 'QTY', 'STATUS']):
                                ws.cell(row=r_idx, column=c).value = "Sold Out" if count == 0 else count
                
                data_row_count += 1

        # 3. AUTO-FIT
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    try: max_len = max(max_len, len(str(cell.value)))
                    except: pass
            ws.column_dimensions[col_letter].width = min((max_len + 2) * 1.1, 50)

    wb.save(output_path)
    print(f"\nSUCCESS! 🚀\nFile saved: {output_path}")

# --- 6. MAIN ---
def main():
    if len(sys.argv) < 3:
        print("Usage: python3 update_v15.py [File1] [File2]")
        return

    path_a = sys.argv[1].strip().replace("'", "").replace('"', "")
    path_b = sys.argv[2].strip().replace("'", "").replace('"', "")

    f1, f2 = os.path.basename(path_a).lower(), os.path.basename(path_b).lower()
    if 'inventory' in f1: inv_path, master_path = path_a, path_b
    elif 'inventory' in f2: inv_path, master_path = path_b, path_a
    else: print("❌ Error: One file must include 'Inventory' in the name."); return

    print(f"\n📂 Inventory: {os.path.basename(inv_path)}")
    print(f"📘 Master Book: {os.path.basename(master_path)}")

    folder = os.path.dirname(master_path)
    output_path = os.path.join(folder, 'Harpeth_Hills_Price_Book_FINAL.xlsx')

    surgical_update(inv_path, master_path, output_path)

if __name__ == "__main__":
    main()
