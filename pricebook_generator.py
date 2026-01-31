"""
Harpeth Hills Dynamic Price Book Generator (First Draft)

Workflow
--------
1) Bootstrap once:
   - Builds price_library.xlsx (editable BasePriceLocked)
   - Fills missing companion crypt baseline using: CompanionCrypt = round_up(...995)(0.8 * 2 * SingleCrypt)
   - Initializes BasePriceLocked = round_up_to_995( baseline_2025 * 1.05 )

2) Publish repeatedly (weekly/monthly):
   - Only replace the FaCTS export file
   - Run publish -> outputs Harpeth_Hills_Price_Book_PUBLISHED.xlsx
   - Uses BasePriceLocked + tiered scarcity; availability comes from FaCTS

Scarcity tiers:
- sold% >= 97% => +20%
- sold% >= 90% => +15%
- else 0%

NOTE: This first draft implements full parsing + scarcity for Mountain View, Building 7/8, Bell Tower.
Building 4/5 are included with simplified availability until we add their detailed mapping logic.
"""

from __future__ import annotations
import re
import math
from pathlib import Path
from typing import Dict, Tuple, List

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ----------------------------
# FILES (fixed names so you only swap FaCTS each time)
# ----------------------------
FACTS_XLSM = Path("Property - Property Inventory with Owner Details.xlsm")  # <- replace this monthly
MASTER_LISTING_XLSX = Path("Harpeth_Hills_Price_Book_REMADE.xlsx")          # <- used only for bootstrap
PRICE_LIBRARY_XLSX = Path("price_library.xlsx")                             # <- editable BasePriceLocked
OUTPUT_XLSX = Path("Harpeth_Hills_Price_Book_PUBLISHED.xlsx")

# ----------------------------
# PRICING RULES
# ----------------------------
ROUND_TO = 995
DEFAULT_INCREASE_PCT = 0.05

# scarcity tiers
TIER1_SOLD = 0.90
TIER1_UPLIFT = 0.15
TIER2_SOLD = 0.97
TIER2_UPLIFT = 0.20

# companion bundle rule (default fill when missing)
COMPANION_DISCOUNT = 0.20  # 20% off two singles


def round_up_to(x: float, base: int = ROUND_TO) -> int:
    """Round up to nearest multiple of base (995)."""
    if x <= 0:
        return 0
    return int(math.ceil(x / base) * base)

def round_up_end995(x: float) -> int:
    """Round up to next price ending in ...995."""
    return int(math.ceil((x + 5) / 1000.0) * 1000 - 5)

def scarcity_uplift(sold_pct: float) -> float:
    """Tiered scarcity uplift."""
    if sold_pct >= TIER2_SOLD:
        return TIER2_UPLIFT
    if sold_pct >= TIER1_SOLD:
        return TIER1_UPLIFT
    return 0.0

def final_price_from_base(base_price_locked: int, sold_pct: float) -> int:
    """Final = BasePriceLocked + tiered scarcity (no compounding)."""
    u = scarcity_uplift(sold_pct)
    if u <= 0:
        return int(base_price_locked)
    return round_up_to(base_price_locked * (1 + u), ROUND_TO)


# ----------------------------
# FORMATTING (baked to match your style)
# ----------------------------
DARK = "404040"
DARKER = "2B2B2B"
MID = "D9D9D9"
GRID = "BFBFBF"
RED = "C00000"

thin = Side(style="thin", color=GRID)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def set_widths(ws, widths: Dict[int, float]):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

def title_bar(ws, title: str, last_col: int):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    c = ws.cell(1, 1, title)
    c.font = Font(bold=True, size=16, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

def header_row(ws, headers: List[str], row: int = 2):
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row, col, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=DARKER)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    ws.row_dimensions[row].height = 18

def group_row(ws, row: int, text: str, last_col: int):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    c = ws.cell(row, 1, text)
    c.font = Font(bold=True)
    c.fill = PatternFill("solid", fgColor=MID)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = BORDER

def write_data_row(ws, row: int, values: List, money_cols: set[int], avail_col: int):
    for col, v in enumerate(values, start=1):
        cell = ws.cell(row, col, v)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="left", vertical="center")

        if col in money_cols and isinstance(v, (int, float)):
            cell.number_format = '"$"#,##0'
            cell.alignment = Alignment(horizontal="right", vertical="center")

    # Availability formatting
    av = ws.cell(row, avail_col)
    av.alignment = Alignment(horizontal="center", vertical="center")
    if isinstance(av.value, (int, float)):
        av.number_format = "0"
    else:
        if str(av.value).strip().lower() == "sold out":
            av.font = Font(color=RED, bold=True)

# ----------------------------
# FaCTS loader
# ----------------------------
def load_facts() -> pd.DataFrame:
    raw = pd.read_excel(FACTS_XLSM, sheet_name=0, engine="openpyxl")
    headers = [re.sub(r"\s+", " ", str(h)).strip() for h in raw.iloc[1].fillna("")]
    df = raw.iloc[2:].copy()
    df.columns = headers
    required_columns = {"Location", "Section", "Status", "Space", "Sales Item"}
    missing = required_columns - set(df.columns)
    if missing:
        missing_list = ", ".join(sorted(missing))
        raise ValueError(f"FaCTS export is missing required columns: {missing_list}")
    df = df[df["Location"].notna()].copy()

    for c in ["Section", "Status", "Type", "Sales Item", "Space", "Right Types"]:
        if c in df.columns:
            df[c] = df[c].astype(str).str.strip()
    return df

def is_available(status: str) -> bool:
    return str(status).strip().lower() == "available"


# ----------------------------
# Sub-product availability parsers
# ----------------------------

# --- Mountain View crypts ---
ROW_THEME_MV = {"D": "D – Touch", "C": "C – Eye", "B": "B – Heart", "A": "A – Prayer"}
MV_CRYPT_RE = re.compile(
    r"Mountain View Mausoleum Crypts\s+(Upper Level|Lower Level)\s+Elevation\s+(\d+)\s+Level\s+([A-D])\s+Crypt\s+(\d+)",
    re.IGNORECASE
)

def count_adjacent_pairs(nums: List[int]) -> int:
    nums = sorted(set(nums))
    used = set()
    pairs = 0
    for n in nums:
        if n in used:
            continue
        if (n + 1) in nums and (n + 1) not in used:
            used.add(n); used.add(n + 1)
            pairs += 1
    return pairs

def mv_buckets(facts: pd.DataFrame) -> Dict[Tuple[str,int,str,str], Dict]:
    """
    Returns dict keyed by (level_band, elevation, row_theme, option) where option in {Single, Companion}
    Each value: {avail, total, sold_pct}
    """
    mv = facts[facts["Section"].str.lower() == "mountain view"].copy()
    rows = []
    for _, r in mv.iterrows():
        m = MV_CRYPT_RE.search(r.get("Space", ""))
        if not m:
            continue
        band = m.group(1).title()
        elev = int(m.group(2))
        row_letter = m.group(3).upper()
        crypt_num = int(m.group(4))
        status = r.get("Status", "")
        sales_item = r.get("Sales Item", "")
        product = "Tandem" if "tandem" in sales_item.lower() else "Single"
        rows.append((band, elev, ROW_THEME_MV.get(row_letter, row_letter), crypt_num, product, status))

    if not rows:
        return {}

    inv = pd.DataFrame(rows, columns=["band","elev","theme","crypt","product","status"])
    out = {}

    for (band, elev, theme), g in inv[inv["product"]=="Single"].groupby(["band","elev","theme"]):
        total = len(g)
        avail = int(g["status"].map(is_available).sum())
        sold_pct = 1 - (avail/total) if total else 0.0
        out[(band,elev,theme,"Single")] = {"avail":avail,"total":total,"sold_pct":sold_pct}

        all_nums = g["crypt"].tolist()
        av_nums = g[g["status"].map(is_available)]["crypt"].tolist()
        total_pairs = count_adjacent_pairs(all_nums)
        avail_pairs = count_adjacent_pairs(av_nums)
        sold_pct_pairs = 1 - (avail_pairs/total_pairs) if total_pairs else 0.0
        out[(band,elev,theme,"Companion")] = {"avail":avail_pairs,"total":total_pairs,"sold_pct":sold_pct_pairs}

    return out


# --- Building 7 / 8 / 5 / Bell Tower / etc. crypt/level patterns ---
# Example from FaCTS:
#   "Last Supper Maus Bldg 7 Crypt/Level 1A"
#   "Bell Tower Mausoleum Crypt/Level 14E-2"
LS_CRYPT_RE = re.compile(
    r"(Last Supper Maus Bldg 7|Last Supper Maus Bldg 8|Last Supper Maus Bldg 5)\s+Crypt/Level\s+([0-9]+)([A-E])(?:-([0-9]+))?",
    re.IGNORECASE
)
BT_CRYPT_RE = re.compile(
    r"(Bell Tower Mausoleum)\s+Crypt/Level\s+([0-9]+)([A-E])(?:-([0-9]+))?",
    re.IGNORECASE
)

ROW_THEME_ABCDE = {
    "E":"E – Heavenly",
    "D":"D (Touch)",
    "C":"C (Eye)",
    "B":"B (Heart)",
    "A":"A (Prayer)",
}

def building_buckets_simple_by_row(
    facts: pd.DataFrame,
    facts_section_name: str,
    space_regex,
    include_tandem: bool = False
) -> Dict[Tuple[str, str], Dict]:
    """
    Computes buckets by row theme only (A-E), for Single and optionally Tandem.
    Key: (row_theme, option)
    Value: {avail,total,sold_pct}

    NOTE: This ignores Covered/Uncovered splits because FaCTS doesn't label them in Space for Bldg 7/8.
    """
    sub = facts[facts["Section"]==facts_section_name].copy()
    rows = []
    for _, r in sub.iterrows():
        s = r.get("Space","")
        m = space_regex.search(s)
        if not m:
            continue
        row_letter = m.group(3).upper()
        status = r.get("Status","")
        sales_item = r.get("Sales Item","")
        is_tandem = "tandem" in sales_item.lower()
        if is_tandem and not include_tandem:
            continue
        opt = "Tandem" if is_tandem else "Single"
        rows.append((ROW_THEME_ABCDE.get(row_letter, row_letter), opt, status))

    if not rows:
        return {}

    inv = pd.DataFrame(rows, columns=["theme","option","status"])
    out = {}
    for (theme, opt), g in inv.groupby(["theme","option"]):
        total = len(g)
        avail = int(g["status"].map(is_available).sum())
        sold_pct = 1 - (avail/total) if total else 0.0
        out[(theme,opt)] = {"avail":avail,"total":total,"sold_pct":sold_pct}
    return out


# ----------------------------
# PRICE LIBRARY bootstrap (includes companion price fill)
# ----------------------------
def bootstrap_price_library():
    if PRICE_LIBRARY_XLSX.exists():
        return

    wb = openpyxl.load_workbook(MASTER_LISTING_XLSX, data_only=True)

    # Extract line items from all “price table” sheets we can detect (Option+Crypt headers)
    skip = set(["README","Availability Dashboard","Sold Out - Reference","Needs Pricing Tables","Cemetery Service Charges","Pricing Issues Tracker"])
    money_re = re.compile(r"\$?\s*([0-9]{1,3}(?:,[0-9]{3})+|[0-9]+)")

    def to_int(v):
        if v is None:
            return None
        if isinstance(v,(int,float)):
            return int(v)
        if isinstance(v,str):
            m = money_re.search(v)
            return int(m.group(1).replace(",","")) if m else None
        return None

    def detect_price_sheets():
        out = []
        for name in wb.sheetnames:
            if name in skip:
                continue
            ws = wb[name]
            found = False
            for r in range(1, 60):
                row = [ws.cell(r,c).value for c in range(1,8)]
                if any(isinstance(v,str) and v.strip().lower()=="option" for v in row) and any(isinstance(v,str) and v.strip().lower()=="crypt" for v in row):
                    found = True
                    break
            if found:
                out.append(name)
        return out

    def parse_sheet(ws, product_family: str):
        # find header row
        header_row = None
        header = {}
        for r in range(1, 120):
            vals = [ws.cell(r,c).value for c in range(1,10)]
            norm = [str(v).strip().lower() if isinstance(v,str) else "" for v in vals]
            if "option" in norm and "crypt" in norm:
                header_row = r
                for c,v in enumerate(norm, start=1):
                    if v:
                        header[v]=c
                break
        if not header_row:
            return []

        opt_c = header.get("option")
        crypt_c = header.get("crypt")
        label_c = header.get("row") or header.get("section") or header.get("product") or header.get("garden")
        front_c = header.get("crypt front") or header.get("plaque(s)") or header.get("niche front") or header.get("laser etching")
        total_c = header.get("total") or header.get("total price") or header.get("all-in total")
        avail_c = header.get("availability") or header.get("# available") or header.get("available")

        # detect plaque prices in header text, if present
        single_plaque = None
        tandem_plaque = None
        companion_plaque = None
        for r in range(1, 25):
            for c in range(1, 8):
                v = ws.cell(r,c).value
                if not isinstance(v,str):
                    continue
                s = v.lower()
                if "single" in s and "plaque" in s:
                    m = money_re.search(v)
                    if m: single_plaque = int(m.group(1).replace(",",""))
                if "tandem" in s and "plaque" in s:
                    m = money_re.search(v)
                    if m: tandem_plaque = int(m.group(1).replace(",",""))
                if "companion" in s and "plaque" in s:
                    m = money_re.search(v)
                    if m: companion_plaque = int(m.group(1).replace(",",""))

        group = None
        theme = None
        recs = []

        for r in range(header_row+1, ws.max_row+1):
            label = ws.cell(r,label_c).value if label_c else None
            opt = ws.cell(r,opt_c).value if opt_c else None
            crypt = ws.cell(r,crypt_c).value if crypt_c else None

            if label is None and opt is None and crypt is None:
                continue

            if isinstance(label,str):
                s = label.strip()
                if s.lower().startswith("elevation") or s.upper() in ["COVERED","UNCOVERED","ALL LEVELS"] or "COVERED" in s.upper() or "UNCOVERED" in s.upper():
                    group = s
                    theme = None
                    continue
                if "–" in s or ("(" in s and any(x in s for x in ["Touch","Eye","Heart","Prayer","Heavenly"])):
                    theme = s

            if not isinstance(opt,str):
                continue
            opt_s = opt.strip()
            if opt_s not in ["Single","Companion","Tandem"]:
                continue

            recs.append({
                "product_family": product_family,
                "group": group,
                "theme": theme,
                "option": opt_s,
                "baseline_2025_crypt": to_int(crypt),
                "baseline_2025_front": to_int(ws.cell(r,front_c).value) if front_c else None,
                "baseline_2025_total": to_int(ws.cell(r,total_c).value) if total_c else None,
                "availability_text": ws.cell(r,avail_c).value if avail_c else None,
                "single_plaque": single_plaque,
                "tandem_plaque": tandem_plaque,
                "companion_plaque": companion_plaque,
            })

        return recs

    price_sheets = detect_price_sheets()
    all_rows = []
    for name in price_sheets:
        all_rows.extend(parse_sheet(wb[name], name))

    df = pd.DataFrame(all_rows)

    # Fill missing companion crypt price from single crypt: 0.8 * (2*single), rounded up to ...995
    single_map = {(r.product_family, r.group, r.theme): r.baseline_2025_crypt
                  for r in df[df["option"]=="Single"].itertuples() if pd.notna(r.baseline_2025_crypt)}
    single_front_map = {(r.product_family, r.group, r.theme): r.baseline_2025_front
                        for r in df[df["option"]=="Single"].itertuples() if pd.notna(r.baseline_2025_front)}

    for idx, r in df.iterrows():
        if r["option"] != "Companion":
            continue
        key = (r["product_family"], r["group"], r["theme"])
        single_crypt = single_map.get(key)
        if pd.isna(r["baseline_2025_crypt"]) and single_crypt is not None:
            df.at[idx,"baseline_2025_crypt"] = round_up_end995(2 * single_crypt * (1 - COMPANION_DISCOUNT))

        # fill companion front (two plaques) if missing
        if pd.isna(r["baseline_2025_front"]):
            if pd.notna(r.get("companion_plaque")):
                df.at[idx,"baseline_2025_front"] = int(r["companion_plaque"])
            elif pd.notna(r.get("single_plaque")):
                df.at[idx,"baseline_2025_front"] = int(r["single_plaque"]) * 2
            else:
                sf = single_front_map.get(key)
                if sf is not None and not pd.isna(sf):
                    df.at[idx,"baseline_2025_front"] = int(sf) * 2

    # fill totals if missing
    m = df["baseline_2025_total"].isna() & df["baseline_2025_crypt"].notna() & df["baseline_2025_front"].notna()
    df.loc[m,"baseline_2025_total"] = df.loc[m,"baseline_2025_crypt"] + df.loc[m,"baseline_2025_front"]

    # initialize BasePriceLocked from baseline * (1+5%), rounded to 995
    df["increase_pct"] = DEFAULT_INCREASE_PCT
    for part in ["crypt","front","total"]:
        src = f"baseline_2025_{part}"
        dst = f"base_price_locked_{part}"
        df[dst] = pd.to_numeric(df[src], errors="coerce").apply(lambda x: round_up_to(x * (1+DEFAULT_INCREASE_PCT), ROUND_TO) if pd.notna(x) else None)

    # policy sheet
    policy = pd.DataFrame([{
        "round_to": ROUND_TO,
        "default_increase_pct": DEFAULT_INCREASE_PCT,
        "companion_discount_pct": COMPANION_DISCOUNT,
        "tier1_sold_pct": TIER1_SOLD,
        "tier1_uplift": TIER1_UPLIFT,
        "tier2_sold_pct": TIER2_SOLD,
        "tier2_uplift": TIER2_UPLIFT,
        "note": "Edit BasePriceLocked_* anytime. Publish runs only update scarcity + availability."
    }])

    with pd.ExcelWriter(PRICE_LIBRARY_XLSX, engine="openpyxl") as w:
        policy.to_excel(w, sheet_name="Pricing_Policy", index=False)
        df.to_excel(w, sheet_name="Price_Library", index=False)

    print("Created price_library.xlsx")


# ----------------------------
# PUBLISH
# ----------------------------
def publish():
    if not PRICE_LIBRARY_XLSX.exists():
        bootstrap_price_library()

    facts = load_facts()
    lib = pd.read_excel(PRICE_LIBRARY_XLSX, sheet_name="Price_Library")

    # Compute FaCTS scarcity buckets
    mv = mv_buckets(facts)
    b7 = building_buckets_simple_by_row(facts, "Last Supper Maus Bldg 7", LS_CRYPT_RE, include_tandem=False)
    b8 = building_buckets_simple_by_row(facts, "Last Supper Maus Bldg 8", LS_CRYPT_RE, include_tandem=False)
    bt = building_buckets_simple_by_row(facts, "Bell Tower Mausoleum", BT_CRYPT_RE, include_tandem=True)

    # Output workbook
    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)

    def build_mausoleum_sheet(product_family: str, title: str, headers: List[str], widths: Dict[int,float]):
        ws = out_wb.create_sheet(product_family[:31])  # excel name limit safeguard
        title_bar(ws, title, last_col=len(headers))
        header_row(ws, headers, row=2)
        ws.freeze_panes = "A3"
        set_widths(ws, widths)

        df = lib[lib["product_family"]==product_family].copy()

        r = 3
        if product_family.startswith("Mountain View"):
            # group by Elevation
            elevs = sorted([e for e in df["group"].dropna().unique() if str(e).lower().startswith("elevation")],
                           key=lambda x: int(re.search(r"(\d+)", str(x)).group(1)))
            theme_order = ["D – Touch","C – Eye","B – Heart","A – Prayer"]

            for g in elevs:
                group_row(ws, r, str(g), last_col=len(headers))
                r += 1
                elev = int(re.search(r"(\d+)", str(g)).group(1))
                band = "Upper Level" if "Upper" in product_family else "Lower Level"

                for theme in theme_order:
                    for opt in ["Single","Companion"]:
                        row_match = df[(df["group"]==g) & (df["theme"]==theme) & (df["option"]==opt)]
                        if len(row_match) != 1:
                            continue
                        rowi = row_match.iloc[0]
                        bucket = mv.get((band,elev,theme,opt), {"avail":0,"total":0,"sold_pct":0.0})
                        sold_pct = float(bucket["sold_pct"])
                        avail = int(bucket["avail"])
                        availability = avail if avail>0 else "Sold Out"

                        crypt = final_price_from_base(int(rowi["base_price_locked_crypt"]), sold_pct) if pd.notna(rowi["base_price_locked_crypt"]) else None
                        front = final_price_from_base(int(rowi["base_price_locked_front"]), sold_pct) if pd.notna(rowi["base_price_locked_front"]) else None
                        total = (crypt + front) if (crypt is not None and front is not None) else None

                        # show theme label only on first line
                        theme_label = theme if opt=="Single" else None

                        write_data_row(ws, r, [theme_label, opt, crypt, front, total, availability],
                                       money_cols={3,4,5}, avail_col=6)
                        r += 1
                r += 1

        else:
            # Building 7/8/BellTower pattern: group labels like COVERED/UNCOVERED exist in library
            groups = [g for g in df["group"].dropna().unique()]
            # keep original order: COVERED after UNCOVERED if present
            def grp_sort(x):
                s=str(x).upper()
                if "UNCOVERED" in s: return (0,s)
                if "COVERED" in s: return (1,s)
                return (2,s)
            groups = sorted(groups, key=grp_sort)

            # theme order for A-E
            theme_order = ["E – Heavenly","D (Touch)","C (Eye)","B (Heart)","A (Prayer)"]

            # choose correct bucket
            bucket_map = None
            if product_family=="Building 7 Mausoleum":
                bucket_map = b7
            elif product_family=="Building 8 Mausoleum":
                bucket_map = b8
            elif product_family=="Bell Tower Mausoleum":
                bucket_map = bt
            else:
                bucket_map = {}

            for g in groups:
                group_row(ws, r, str(g), last_col=len(headers))
                r += 1

                for theme in theme_order:
                    for opt in ["Single","Tandem","Companion"]:
                        row_match = df[(df["group"]==g) & (df["theme"]==theme) & (df["option"]==opt)]
                        if len(row_match) != 1:
                            continue
                        rowi = row_match.iloc[0]

                        # sold_pct by row theme+opt (Companion uses Single sold% as proxy in this first draft)
                        if opt == "Companion":
                            b = bucket_map.get((theme,"Single"), {"avail":0,"total":0,"sold_pct":0.0})
                            # availability of companion units should be computed later; for now, if singles sold out -> companion sold out.
                            sold_pct = float(b["sold_pct"])
                            avail = int(b["avail"])
                            availability = avail if avail>0 else "Sold Out"
                        else:
                            b = bucket_map.get((theme,opt), {"avail":0,"total":0,"sold_pct":0.0})
                            sold_pct = float(b["sold_pct"])
                            avail = int(b["avail"])
                            availability = avail if avail>0 else "Sold Out"

                        crypt = final_price_from_base(int(rowi["base_price_locked_crypt"]), sold_pct) if pd.notna(rowi["base_price_locked_crypt"]) else None
                        front = final_price_from_base(int(rowi["base_price_locked_front"]), sold_pct) if pd.notna(rowi["base_price_locked_front"]) else None
                        total = (crypt + front) if (crypt is not None and front is not None) else None

                        theme_label = theme if opt=="Single" else None
                        write_data_row(ws, r, [theme_label, opt, crypt, front, total, availability],
                                       money_cols={3,4,5}, avail_col=6)
                        r += 1
                r += 1

        return ws

    # Build sheets (first draft coverage)
    build_mausoleum_sheet("Mountain View - Upper Level",
                          "MOUNTAIN VIEW MAUSOLEUM — UPPER LEVEL",
                          ["Row","Option","Crypt","Crypt Front","Total","Availability"],
                          {1:22,2:12,3:14,4:14,5:14,6:12})

    build_mausoleum_sheet("Mountain View - Lower Level",
                          "MOUNTAIN VIEW MAUSOLEUM — LOWER LEVEL",
                          ["Row","Option","Crypt","Crypt Front","Total","Availability"],
                          {1:22,2:12,3:14,4:14,5:14,6:12})

    build_mausoleum_sheet("Building 7 Mausoleum",
                          "BUILDING 7 MAUSOLEUM",
                          ["Row","Option","Crypt","Crypt Front","Total","Availability"],
                          {1:22,2:12,3:14,4:14,5:14,6:12})

    build_mausoleum_sheet("Building 8 Mausoleum",
                          "BUILDING 8 MAUSOLEUM",
                          ["Section","Option","Crypt","Crypt Front","Total","Availability"],
                          {1:28,2:12,3:14,4:14,5:14,6:12})

    build_mausoleum_sheet("Bell Tower Mausoleum",
                          "BELL TOWER MAUSOLEUM",
                          ["Row","Option","Crypt","Crypt Front","Total","Availability"],
                          {1:22,2:12,3:14,4:14,5:14,6:12})

    out_wb.save(OUTPUT_XLSX)
    print(f"Wrote: {OUTPUT_XLSX}")


# ----------------------------
# CLI
# ----------------------------
if __name__ == "__main__":
    import sys
    cmd = sys.argv[1].lower() if len(sys.argv)>1 else "publish"
    if cmd == "bootstrap":
        bootstrap_price_library()
        print("Bootstrap complete. Edit BasePriceLocked in price_library.xlsx as needed.")
    elif cmd == "publish":
        publish()
    else:
        raise SystemExit("Usage: python pricebook_generator.py bootstrap|publish")
